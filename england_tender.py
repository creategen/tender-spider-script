#!/usr/bin/env python3
"""
英国 Contracts Finder 招标数据爬取脚本
从 https://www.contractsfinder.service.gov.uk/Search/Results 爬取全部招标信息

策略：
  - 列表页：Playwright 自动化浏览器（处理 JS 渲染 + 翻页）
  - 详情页：requests + BeautifulSoup（直接 HTTP 请求，避免 DOM 引用失效）
  - 参考 detail.py 的解析逻辑

功能：
  - 翻页爬取全部数据
  - 提取详情页所有字段（URL、Title、Department、Statement、Contract summary、
    Description、More information、How to apply、About the buyer）
  - 保存到 Excel，支持一级/二级表头（合并单元格）
  - 每 10000 条数据保存一次，避免单文件过大
  - 文件保存到桌面"英国招标"文件夹
  - 上传至 Gofile.io 并发送下载链接邮件

用法:
# 默认：爬取前2页
python england_tender.py

# 爬取前5页
python england_tender.py --max-pages 5

# 爬取全部页
python england_tender.py --max-pages full

# 带邮件发送（爬取前2页）
python england_tender.py --sender xxx@qq.com --auth-code xxx --receiver xxx@qq.com --start-page 1 --max-pages 2
"""

import argparse
import os
import re
import smtplib
import sys
import time
from datetime import datetime, timezone, timedelta
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from urllib3.exceptions import InsecureRequestWarning
import urllib3

import requests
from bs4 import BeautifulSoup
from playwright.sync_api import sync_playwright
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# 禁用 SSL 警告（开发/测试阶段）
urllib3.disable_warnings(InsecureRequestWarning)

# ==================== 固定配置 ====================

TARGET_URL = "https://www.contractsfinder.service.gov.uk/Search/Results"
BASE_URL = "https://www.contractsfinder.service.gov.uk"
LIST_SELECTOR = "#dashboard_notices > div.gadget-body"
# 等待搜索结果出现的选择器（更精确地等待内容加载完成）
SEARCH_RESULT_SELECTOR = "#dashboard_notices .search-result"
NEXT_PAGE_SELECTOR = (
    "#dashboard_notices > div.gadget-footer > ul > "
    "li.standard-paginate-next-box.standard-paginate-next-icon"
)
VIEWPORT = {"width": 1280, "height": 800}
USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/120.0.0.0 Safari/537.36"
)
BATCH_SIZE = 10000  # 每 10000 条保存一次
REQUEST_DELAY = 1.5  # 每次详情页请求间隔（秒），防止 429
MAX_RETRIES = 3      # 429 重试最大次数
RETRY_DELAYS = [121, 121, 121]  # 每次重试的等待时间（秒），递增等待策略

# 保存目录
SAVE_DIR = os.path.join(os.getcwd(), "英国招标")

# QQ邮箱附件限制
QQ_MAX_SINGLE_FILE_SIZE = 20 * 1024 * 1024  # 20MB
QQ_MAX_TOTAL_SIZE = 50 * 1024 * 1024  # 50MB

# Gofile API
GOFILE_API = "https://api.gofile.io"


# ==================== HTTP 会话（用于详情页请求）====================

http_session = requests.Session()
http_session.headers.update({"User-Agent": USER_AGENT})
http_session.verify = False  # 开发环境关闭 SSL 验证

# ==================== 表头定义 ====================
# 一级表头（占两行的字段）和二级表头的结构
# 格式: (一级表头名, [二级表头列表])  如果二级表头为空列表，表示该字段占两行

HEADER_STRUCTURE = [
    ("URL", []),
    ("Title", []),
    ("Department", []),
    ("Statement", []),
    ("Contract summary", [
        "Industry",
        "Location of contract",
        "Value of contract",
        "Procurement reference",
        "Published date",
        "Closing date",
        "Closing time",
        "Contract start date",
        "Contract end date",
        "Contract type",
        "Procedure type",
        "Contract is suitable for SMEs?",
        "Contract is suitable for VCSEs?",
    ]),
    ("Description", []),
    ("More information", [
        "Links",
        "Tender notice",
        "NOTE",
    ]),
    ("How to apply", []),
    ("About the buyer", [
        "Contact name",
        "Address",
        "Telephone",
        "Email",
        "Website",
    ]),
]

# 扁平化的二级表头列表（用于数据字典的 key 和 Excel 第二行表头）
FLAT_HEADERS = []
for parent, children in HEADER_STRUCTURE:
    if children:
        for child in children:
            FLAT_HEADERS.append(child)
    else:
        FLAT_HEADERS.append(parent)


# ==================== Excel 写入 ====================

def save_to_excel(batch_data, file_index):
    """
    将一批数据保存到 Excel 文件，使用合并单元格实现一级/二级表头。

    表头结构：
      第 1 行：一级表头（有二级表头的合并到对应列范围，无二级表头的合并两行）
      第 2 行：二级表头（无二级表头的列与第 1 行合并）
      第 3 行起：数据
    """
    if not batch_data:
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "招标数据"

    # --- 样式定义 ---
    header_font = Font(name="Arial", bold=True, size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    data_alignment = Alignment(vertical="top", wrap_text=True)

    # --- 写入一级表头（第 1 行）和二级表头（第 2 行）---
    col = 1  # 当前列号（1-based）
    for parent, children in HEADER_STRUCTURE:
        if children:
            # 有二级表头：一级表头合并到对应列范围（第 1 行）
            start_col = col
            end_col = col + len(children) - 1
            ws.merge_cells(start_row=1, start_column=start_col,
                           end_row=1, end_column=end_col)
            cell = ws.cell(row=1, column=start_col, value=parent)
            cell.font = header_font_white
            cell.alignment = header_alignment
            cell.fill = header_fill
            cell.border = thin_border
            # 填充合并区域其他单元格的边框
            for c in range(start_col, end_col + 1):
                ws.cell(row=1, column=c).border = thin_border
                ws.cell(row=1, column=c).fill = header_fill

            # 二级表头写在第 2 行
            for i, child in enumerate(children):
                cell = ws.cell(row=2, column=start_col + i, value=child)
                cell.font = header_font
                cell.alignment = header_alignment
                cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
                cell.border = thin_border

            col = end_col + 1
        else:
            # 无二级表头：一级表头合并第 1 行和第 2 行
            ws.merge_cells(start_row=1, start_column=col,
                           end_row=2, end_column=col)
            cell = ws.cell(row=1, column=col, value=parent)
            cell.font = header_font_white
            cell.alignment = header_alignment
            cell.fill = header_fill
            cell.border = thin_border
            ws.cell(row=2, column=col).border = thin_border
            ws.cell(row=2, column=col).fill = header_fill
            col += 1

    # --- 写入数据（从第 3 行开始）---
    for row_idx, record in enumerate(batch_data, start=3):
        for col_idx, key in enumerate(FLAT_HEADERS, start=1):
            value = record.get(key, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = data_alignment
            cell.border = thin_border

    # --- 自动调整列宽 ---
    for col_idx in range(1, len(FLAT_HEADERS) + 1):
        max_length = 0
        col_letter = get_column_letter(col_idx)
        for row in ws.iter_rows(min_row=1, max_row=min(len(batch_data) + 2, 100),
                                min_col=col_idx, max_col=col_idx):
            for cell in row:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
        adjusted_width = min(max_length + 2, 60)
        ws.column_dimensions[col_letter].width = max(adjusted_width, 10)

    # --- 冻结窗格（冻结表头两行）---
    ws.freeze_panes = "A3"

    # --- 保存文件 ---
    # 使用东八区时间
    tz_beijing = timezone(timedelta(hours=8))
    timestamp = datetime.now(tz_beijing).strftime("%Y-%m-%d-%H-%M-%S")
    os.makedirs(SAVE_DIR, exist_ok=True)
    file_name = f"英国-招标-{timestamp}.xlsx"
    file_path = os.path.join(SAVE_DIR, file_name)

    try:
        wb.save(file_path)
        print(f"\n[保存] 已保存 {len(batch_data)} 条数据到 {file_path}\n")
        return file_path
    except Exception as e:
        print(f"[错误] 保存文件失败: {e}")
        return None


# ==================== 详情页数据提取（requests + BeautifulSoup）====================

def fetch_detail_page(url):
    """
    用 requests 获取详情页 HTML 并解析为 BeautifulSoup 对象。
    遇到 429 Too Many Requests 时按 RETRY_DELAYS 指定的间隔重试。
    """
    for attempt in range(MAX_RETRIES):
        try:
            response = http_session.get(url, timeout=30)
            response.raise_for_status()
            break
        except requests.exceptions.HTTPError as e:
            if response.status_code == 429:
                wait_time = RETRY_DELAYS[min(attempt, len(RETRY_DELAYS) - 1)]
                print(f"    [429] 请求过快，第 {attempt + 1}/{MAX_RETRIES-1} 次重试，等待 {wait_time}s...")
                time.sleep(wait_time)
                if attempt == MAX_RETRIES - 1:
                    raise
            else:
                raise
        except requests.exceptions.ConnectionError:
            wait_time = RETRY_DELAYS[min(attempt, len(RETRY_DELAYS) - 1)]
            print(f"    [连接错误] 第 {attempt + 1}/{MAX_RETRIES-1} 次重试，等待 {wait_time}s...")
            time.sleep(wait_time)
            if attempt == MAX_RETRIES - 1:
                raise
        except requests.exceptions.Timeout:
            wait_time = RETRY_DELAYS[min(attempt, len(RETRY_DELAYS) - 1)]
            print(f"    [超时] 第 {attempt + 1}/{MAX_RETRIES-1} 次重试，等待 {wait_time}s...")
            time.sleep(wait_time)
            if attempt == MAX_RETRIES - 1:
                raise

    soup = BeautifulSoup(response.text, "html.parser")
    # 将 <br> 替换为空格（与 detail.py 一致）
    for br in soup.find_all("br"):
        br.replace_with("  ")
    return soup


def extract_detail(url):
    """
    从详情页提取所有目标字段。
    使用 requests + BeautifulSoup，参考 detail.py 的解析逻辑。
    """
    data = {key: "" for key in FLAT_HEADERS}
    data["URL"] = url

    try:
        soup = fetch_detail_page(url)

        # --- Title ---
        title_elem = soup.select_one("#all-content-wrapper > h1")
        if title_elem:
            data["Title"] = title_elem.get_text(strip=True)

        # --- Department ---
        dept_elem = soup.select_one("#home-breadcrumb-description > h2")
        if dept_elem:
            data["Department"] = dept_elem.get_text(strip=True)

        # --- Statement ---
        statement_elem = soup.select_one("#content-block-left > p")
        if statement_elem:
            data["Statement"] = statement_elem.get_text(strip=True)

        # 获取所有 content-block 区块
        content_blocks = soup.find_all("div", class_="content-block")

        # --- Contract summary ---
        contract_summary_block = None
        for block in content_blocks:
            h3 = block.find("h3")
            if h3 and "Contract summary" in h3.get_text():
                contract_summary_block = block
                break

        if contract_summary_block:
            contract_field_mapping = {
                "Industry": "Industry",
                "Location of contract": "Location of contract",
                "Value of contract": "Value of contract",
                "Procurement reference": "Procurement reference",
                "Published date": "Published date",
                "Closing date": "Closing date",
                "Closing time": "Closing time",
                "Contract start date": "Contract start date",
                "Contract end date": "Contract end date",
                "Contract type": "Contract type",
                "Procedure type": "Procedure type",
                "Contract is suitable for SMEs?": "Contract is suitable for SMEs?",
                "Contract is suitable for VCSEs?": "Contract is suitable for VCSEs?",
            }
            h4_elements = contract_summary_block.find_all("h4")
            for h4 in h4_elements:
                label = h4.get_text(strip=True)
                next_elem = h4.find_next_sibling()
                value = ""
                if next_elem:
                    if next_elem.name == "p":
                        value = next_elem.get_text(strip=True)
                    elif next_elem.name == "ul":
                        items = [li.get_text(strip=True) for li in next_elem.find_all("li")]
                        value = "; ".join(filter(None, items))
                for key, mapped_key in contract_field_mapping.items():
                    if key in label:
                        data[mapped_key] = value
                        break

        # --- Description ---
        for block in content_blocks:
            h3 = block.find("h3")
            if h3 and "Description" in h3.get_text():
                paragraphs = [p.get_text(strip=True) for p in block.find_all("p")]
                data["Description"] = "".join(filter(None, paragraphs))
                break

        # --- More information ---
        more_info_block = None
        for block in content_blocks:
            h3 = block.find("h3")
            if h3 and "More information" in h3.get_text():
                more_info_block = block
                break

        if more_info_block:
            # Links: 提取所有链接
            links = []
            for a in more_info_block.find_all("a", href=True):
                href = a["href"]
                text = a.get_text(strip=True)
                if href.startswith("http"):
                    links.append(f"{text}: {href}")
            if links:
                data["Links"] = "\n".join(links)

            # Tender notice 和 NOTE: 通过 h4 查找
            more_info_field_mapping = {
                "Tender notice": "Tender notice",
                "NOTE": "NOTE",
            }
            h4_elements = more_info_block.find_all("h4")
            for h4 in h4_elements:
                label = h4.get_text(strip=True)
                next_elem = h4.find_next_sibling()
                value = ""
                if next_elem:
                    if next_elem.name == "p":
                        value = next_elem.get_text(strip=True)
                    elif next_elem.name == "ul":
                        items = [li.get_text(strip=True) for li in next_elem.find_all("li")]
                        value = "; ".join(filter(None, items))
                for key, mapped_key in more_info_field_mapping.items():
                    if key in label:
                        data[mapped_key] = value
                        break

        # --- How to apply ---
        for block in content_blocks:
            h3 = block.find("h3")
            if h3 and "How to apply" in h3.get_text():
                paragraphs = [p.get_text(strip=True) for p in block.find_all("p")]
                data["How to apply"] = "\n".join(filter(None, paragraphs))
                break

        # --- About the buyer ---
        buyer_block = None
        for block in content_blocks:
            h3 = block.find("h3")
            if h3 and "About the buyer" in h3.get_text():
                buyer_block = block
                break

        if buyer_block:
            buyer_field_mapping = {
                "Contact name": "Contact name",
                "Address": "Address",
                "Telephone": "Telephone",
                "Email": "Email",
                "Website": "Website",
            }
            h4_elements = buyer_block.find_all("h4")
            for h4 in h4_elements:
                label = h4.get_text(strip=True)
                next_elem = h4.find_next_sibling()
                value = ""
                if next_elem:
                    if next_elem.name == "p":
                        value = next_elem.get_text(strip=False)
                    elif next_elem.name == "a":
                        value = next_elem.get_text(strip=True)
                for key, mapped_key in buyer_field_mapping.items():
                    if key in label:
                        data[mapped_key] = value
                        break

    except Exception as e:
        print(f"  [警告] 详情页解析出错 ({url}): {e}")

    return data


# ==================== 列表页链接收集 ====================

def collect_page_links(page, max_retries=3):
    """
    从当前列表页收集所有文章的详情页 URL。
    仅收集链接，不点击进入详情页，避免 DOM 引用失效。
    增加重试机制处理页面加载超时问题。
    """
    retry_count = 0
    last_error = None

    while retry_count < max_retries:
        try:
            # 等待搜索结果出现（比等容器更可靠，确保内容已加载）
            try:
                page.wait_for_selector(SEARCH_RESULT_SELECTOR, timeout=60000)
            except Exception:
                # 回退：等待容器出现
                page.wait_for_selector(LIST_SELECTOR, timeout=30000)

            # 额外等待确保内容完全加载
            page.wait_for_timeout(3000)

            # 用 JavaScript 一次性提取所有链接，避免逐个查询 DOM 元素
            links = page.evaluate("""() => {
                const items = document.querySelectorAll('#dashboard_notices .search-result h2 a');
                return Array.from(items).map(a => a.href);
            }""")

            # 如果没有找到链接，可能是页面还没完全加载，重试
            if not links:
                retry_count += 1
                if retry_count < max_retries:
                    print(f"  [警告] 未找到链接（可能页面未完全加载），第 {retry_count}/{max_retries} 次重试...")
                    page.wait_for_timeout(5000)
                    continue
                else:
                    print(f"  [错误] 多次重试仍未找到链接")
                    return []

            # 过滤并补全 URL
            result = []
            for href in links:
                if not href:
                    continue
                if href.startswith("/"):
                    href = BASE_URL + href
                result.append(href)

            return result

        except Exception as e:
            last_error = e
            retry_count += 1
            if retry_count < max_retries:
                print(f"  [警告] 收集链接失败，第 {retry_count}/{max_retries} 次重试...")
                page.wait_for_timeout(5000)
            else:
                print(f"  [错误] 收集链接失败，已重试 {max_retries} 次: {e}")
                raise

    raise last_error


# ==================== Gofile 上传 ====================

def create_gofile_account():
    """创建 Gofile 匿名账户，返回 token 和 rootFolderId"""
    resp = requests.post(f"{GOFILE_API}/accounts")
    data = resp.json()["data"]
    return data["token"], data["rootFolder"]


def upload_to_gofile(filepath, token, folder_id, max_retries=2):
    """上传单个文件到 Gofile，支持重试，返回文件信息字典"""
    filename = os.path.basename(filepath)
    filesize = os.path.getsize(filepath)
    if filesize >= 1024 * 1024:  # >= 1MB
        size_str = f"{filesize / 1024 / 1024:.1f} MB"
    else:
        size_str = f"{filesize / 1024:.1f} KB"
    print(f"  上传: {filename} ({size_str})")

    last_error = None
    for attempt in range(max_retries + 1):
        try:
            if attempt > 0:
                print(f"  第 {attempt + 1} 次重试上传...")

            # 获取最佳服务器
            resp = requests.get(f"{GOFILE_API}/servers", params={"token": token})
            servers = resp.json()["data"]["servers"]
            na_servers = [s["name"] for s in servers if s["zone"] == "na"]
            server = na_servers[0] if na_servers else servers[0]["name"]

            # 上传
            upload_url = f"https://{server}.gofile.io/contents/uploadfile"
            start_time = time.time()
            with open(filepath, "rb") as f:
                files = {"file": (filename, f)}
                data = {"token": token, "folderId": folder_id}
                resp = requests.post(upload_url, files=files, data=data, timeout=1800)

            elapsed = time.time() - start_time
            result = resp.json()
            if result.get("status") != "ok":
                raise Exception(f"上传失败: {result}")

            download_page = result["data"]["downloadPage"]
            print(f"  上传成功! ({elapsed:.1f}s) 下载页: {download_page}")

            return {
                "filename": filename,
                "downloadPage": download_page,
                "fileId": result["data"]["id"],
                "size": filesize,
            }
        except Exception as e:
            last_error = e
            print(f"  上传失败(尝试 {attempt + 1}/{max_retries + 1}): {e}")
            if attempt < max_retries:
                time.sleep(3)  # 重试前等待3秒

    raise Exception(f"上传失败，已重试 {max_retries} 次: {last_error}")


def set_gofile_folder_public(folder_id, token):
    """设置 Gofile 文件夹为公开访问（关键步骤！）"""
    resp = requests.put(
        f"{GOFILE_API}/contents/{folder_id}/update",
        headers={
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json",
        },
        json={"attribute": "public", "attributeValue": True},
    )
    return resp.json()


def upload_to_gofile(file_paths):
    """
    将多个文件上传到 Gofile 同一目录，返回下载链接和上传结果
    使用新的辅助函数，支持重试机制
    """
    print("\n[步骤] 上传文件到 Gofile.io...")

    try:
        token, root_folder = create_gofile_account()
        print(f"  Gofile 账户已创建")
    except Exception as e:
        print(f"  创建 Gofile 账户失败: {e}")
        return None, []

    # 按文件大小排序，小文件先上传
    files_with_size = [(fp, os.path.getsize(fp)) for fp in file_paths]
    files_sorted = sorted(files_with_size, key=lambda x: x[1])

    upload_results = []
    for filepath, filesize in files_sorted:
        try:
            result = upload_to_gofile(filepath, token, root_folder)
            upload_results.append(result)
        except Exception as e:
            print(f"  上传失败: {os.path.basename(filepath)} - {e}")

    if not upload_results:
        print("  所有文件上传失败")
        return None, []

    # 关键：设置文件夹公开访问
    print("\n  设置文件夹为公开访问...")
    try:
        set_gofile_folder_public(root_folder, token)
        print("  文件夹已设为公开")
    except Exception as e:
        print(f"  设置公开失败: {e}")

    download_link = upload_results[0]["downloadPage"]
    print(f"  下载链接: {download_link}")
    return download_link, upload_results


# ==================== 邮件发送 ====================

def send_email_with_gofile_link(subject, sender, auth_code, recipient, download_link, upload_results):
    """发送包含Gofile下载链接的邮件（HTML格式）"""
    print(f"\n发送邮件到 {recipient}...")

    # 文件清单表格行
    file_rows_html = ""
    file_rows_text = ""
    for i, r in enumerate(upload_results, 1):
        bg = "#f2f2f2" if i % 2 == 1 else "#ffffff"
        size_bytes = r['size']
        if size_bytes >= 1024 * 1024:  # >= 1MB
            size_str = f"{size_bytes/1024/1024:.1f} MB"
        else:
            size_str = f"{size_bytes/1024:.1f} KB"
        file_rows_html += f"""<tr style="background-color: {bg};">
<td style="border: 1px solid #ddd; padding: 10px;">{i}</td>
<td style="border: 1px solid #ddd; padding: 10px;">{r['filename']}</td>
<td style="border: 1px solid #ddd; padding: 10px;">{size_str}</td></tr>
"""
        file_rows_text += f"{i}. {r['filename']} ({size_str})\n"

    now_str = datetime.now(timezone(timedelta(hours=8))).strftime('%Y年%m月%d日')

    html_body = f"""<html><body style="font-family: Microsoft YaHei, Arial, sans-serif; color: #333; line-height: 1.8; max-width: 700px; margin: 0 auto;">
<h2 style="color: #1a5276; border-bottom: 2px solid #2980b9; padding-bottom: 8px;">英国 Contracts Finder 招标数据爬取结果</h2>
<div style="background-color: #eaf2f8; padding: 15px; border-radius: 5px; margin: 15px 0;">
<strong>爬取时间：</strong>{now_str}<br>
<strong>数据来源：</strong>https://www.contractsfinder.service.gov.uk<br>
<strong>总计：</strong>{len(upload_results)}个文件
</div>
<h3>下载链接</h3>
<div style="background-color: #fff3cd; border: 1px solid #ffc107; padding: 20px; border-radius: 5px; text-align: center; margin: 15px 0;">
<a href="{download_link}" style="font-size: 18px; color: #2980b9; font-weight: bold; text-decoration: none;">点击下载全部文件</a>
<br><span style="color: #888; font-size: 13px;">{download_link}</span>
<br><span style="color: #666; font-size: 12px;">（{len(upload_results)}个文件均在同一目录下，可逐个下载）</span>
</div>
<h3>文件清单</h3>
<table style="border-collapse: collapse; width: 100%;">
<tr style="background-color: #2980b9; color: white;"><th style="padding: 10px; text-align: left;">序号</th><th style="padding: 10px; text-align: left;">文件名</th><th style="padding: 10px; text-align: left;">大小</th></tr>
{file_rows_html}
</table>
<div style="background-color: #f8d7da; border: 1px solid #f5c6cb; padding: 15px; border-radius: 5px; margin: 20px 0;">
<strong style="color: #721c24;">提醒：Gofile.io 链接有效期有限，建议尽快下载。</strong>
</div>
<p style="color:#888; font-size:12px; margin-top:20px;">此邮件由自动化爬取系统发送，请勿直接回复。</p>
</body></html>"""

    text_body = f"""英国 Contracts Finder 招标数据爬取结果

爬取时间：{now_str}
数据来源：https://www.contractsfinder.service.gov.uk
总计：{len(upload_results)}个文件

下载链接（{len(upload_results)}个文件在同一目录下，可逐个下载）：
{download_link}

文件清单：
{file_rows_text}
提醒：Gofile.io 链接有效期有限，建议尽快下载。
"""

    msg = MIMEMultipart("alternative")
    msg["From"] = sender
    msg["To"] = recipient
    msg["Subject"] = subject

    msg.attach(MIMEText(text_body, "plain", "utf-8"))
    msg.attach(MIMEText(html_body, "html", "utf-8"))

    try:
        server = smtplib.SMTP_SSL("smtp.qq.com", 465)
        server.login(sender, auth_code)
        server.sendmail(sender, recipient, msg.as_string())
        server.quit()
        print("  [OK] 邮件发送成功！")
        return True
    except Exception as e:
        # 尝试TLS方式
        try:
            server = smtplib.SMTP("smtp.qq.com", 587)
            server.starttls()
            server.login(sender, auth_code)
            server.sendmail(sender, recipient, msg.as_string())
            server.quit()
            print("  [OK] 邮件发送成功(TLS)！")
            return True
        except Exception as e2:
            print(f"  [FAIL] 邮件发送失败: {e2}")
            return False


def send_email_with_attachments(subject, sender, auth_code, recipient, files):
    """发送带附件的邮件（符合QQ邮箱规则）"""
    print(f"\n发送邮件到 {recipient}（带附件）...")

    # 文件清单
    file_rows_text = ""
    total_size = 0
    for i, (filepath, filename, filesize) in enumerate(files, 1):
        total_size += filesize
        if filesize >= 1024 * 1024:  # >= 1MB
            size_str = f"{filesize/1024/1024:.1f} MB"
        else:
            size_str = f"{filesize/1024:.1f} KB"
        file_rows_text += f"{i}. {filename} ({size_str})\n"

    now_str = datetime.now(timezone(timedelta(hours=8))).strftime('%Y年%m月%d日')

    html_body = f"""<html><body style="font-family: Microsoft YaHei, Arial, sans-serif; color: #333; line-height: 1.8; max-width: 700px; margin: 0 auto;">
<h2 style="color: #1a5276; border-bottom: 2px solid #2980b9; padding-bottom: 8px;">英国 Contracts Finder 招标数据爬取结果</h2>
<div style="background-color: #eaf2f8; padding: 15px; border-radius: 5px; margin: 15px 0;">
<strong>爬取时间：</strong>{now_str}<br>
<strong>数据来源：</strong>https://www.contractsfinder.service.gov.uk<br>
<strong>总计：</strong>{len(files)}个文件（作为附件发送）
</div>
<h3>文件清单</h3>
<pre style="background-color: #f5f5f5; padding: 15px; border-radius: 5px; font-family: Consolas, Monaco, monospace;">{file_rows_text}</pre>
<div style="background-color: #d4edda; border: 1px solid #c3e6cb; padding: 15px; border-radius: 5px; margin: 20px 0;">
<strong style="color: #155724;">✓ 文件已作为附件发送，可直接下载。</strong>
</div>
<p style="color:#888; font-size:12px; margin-top:20px;">此邮件由自动化爬取系统发送，请勿直接回复。</p>
</body></html>"""

    text_body = f"""英国 Contracts Finder 招标数据爬取结果

爬取时间：{now_str}
数据来源：https://www.contractsfinder.service.gov.uk
总计：{len(files)}个文件（作为附件发送）

文件清单：
{file_rows_text}
文件已作为附件发送，可直接下载。
"""

    msg = MIMEMultipart("mixed")
    msg["From"] = sender
    msg["To"] = recipient
    msg["Subject"] = subject

    msg.attach(MIMEText(text_body, "plain", "utf-8"))
    msg.attach(MIMEText(html_body, "html", "utf-8"))

    # 添加附件
    for filepath, filename, filesize in files:
        try:
            with open(filepath, "rb") as f:
                part = MIMEBase("application", "octet-stream")
                part.set_payload(f.read())
                encoders.encode_base64(part)
                # 正确设置附件文件名
                part.add_header('Content-Disposition', 'attachment', filename=('utf-8', '', filename))
                msg.attach(part)
            print(f"  已添加附件: {filename}")
        except Exception as e:
            print(f"  添加附件失败: {filename} - {e}")

    try:
        server = smtplib.SMTP_SSL("smtp.qq.com", 465)
        server.login(sender, auth_code)
        server.sendmail(sender, recipient, msg.as_string())
        server.quit()
        print("  [OK] 邮件发送成功（带附件）！")
        return True
    except Exception as e:
        # 尝试TLS方式
        try:
            server = smtplib.SMTP("smtp.qq.com", 587)
            server.starttls()
            server.login(sender, auth_code)
            server.sendmail(sender, recipient, msg.as_string())
            server.quit()
            print("  [OK] 邮件发送成功（带附件，TLS）！")
            return True
        except Exception as e2:
            print(f"  [FAIL] 邮件发送失败: {e2}")
            return False


# ==================== 主流程 ====================

def main(max_pages=2, start_page=1, sender=None, auth_code=None, receiver=None):
    """
    主流程。
    max_pages: 最大爬取页数，默认 2 页。
               从 start_page 开始计算，爬取 max_pages 页。
               可以是数字 n（从 start_page 开始爬取 n 页）或 "full"（爬取全部页）。
    start_page: 起始页码，默认 1。
    sender/auth_code/receiver: 邮件参数，全部非空时上传并发送邮件。
    """
    print("=" * 60)
    print("英国 Contracts Finder 招标数据爬取")
    print(f"目标网址: {TARGET_URL}")
    print(f"保存目录: {SAVE_DIR}")
    print(f"分批保存: 每 {BATCH_SIZE} 条")
    if start_page > 1:
        print(f"起始页码: 第 {start_page} 页")
    if max_pages == "full":
        print("模式: 爬取全部页数据")
    elif max_pages > 0:
        end_page = start_page + max_pages - 1
        print(f"爬取范围: 第 {start_page} 页到第 {end_page} 页（共 {max_pages} 页）")
    else:
        print(f"默认模式: 从第 {start_page} 页开始爬取")
    print("=" * 60)

    os.makedirs(SAVE_DIR, exist_ok=True)

    batch_data = []     # 当前批次数据（待保存）
    counter = 0         # 总计数
    file_index = 0      # 文件序号
    saved_files = []    # 所有已保存的 Excel 文件路径
    saved_records = []  # 每个文件对应的记录数
    
    # 计算结束页码
    end_page = None
    if max_pages != "full" and max_pages > 0:
        end_page = start_page + max_pages - 1

    with sync_playwright() as p:
        # 启动浏览器
        print("\n[步骤1] 初始化浏览器...")
        browser = p.chromium.launch(
            headless=True,
            args=["--no-sandbox", "--disable-setuid-sandbox", "--disable-dev-shm-usage"],
        )
        context = browser.new_context(
            viewport=VIEWPORT,
            user_agent=USER_AGENT,
            bypass_csp=True,
        )
        page = context.new_page()
        page.set_default_timeout(120000)
        page.set_default_navigation_timeout(120000)

        # 访问目标页面
        print("[步骤2] 访问目标页面...")
        page.goto(TARGET_URL, wait_until="domcontentloaded", timeout=120000)
        page.wait_for_load_state("load", timeout=60000)
        print("  页面加载完成")

        # 如果指定了起始页码大于1，则跳转到指定页
        if start_page > 1:
            print(f"  跳转到第 {start_page} 页...")
            jump_url = f"{TARGET_URL}?&page={start_page}#dashboard_notices"
            try:
                page.goto(jump_url, wait_until="domcontentloaded", timeout=120000)
                page.wait_for_selector(LIST_SELECTOR, timeout=60000)
                print(f"  已跳转到第 {start_page} 页")
            except Exception as e:
                print(f"  [警告] 跳转到第 {start_page} 页失败: {e}")
                print("  将从第 1 页开始爬取")
                start_page = 1
                page_num = 1
            else:
                page_num = start_page
        else:
            page_num = 1
        while True:
            print(f"\n--- 正在处理第 {page_num} 页 ---")

            # 收集当前页所有详情页链接
            try:
                detail_urls = collect_page_links(page)
                print(f"  本页找到 {len(detail_urls)} 个列表项")
            except Exception as e:
                print(f"  [错误] 第 {page_num} 页收集链接失败: {e}")
                print(f"  [提示] 可能已到达最后一页或页面结构发生变化")
                break

            # 如果本页没有找到链接，可能是页面结构变化或已结束
            if not detail_urls:
                print(f"  [警告] 第 {page_num} 页未找到任何链接，爬取结束。")
                break

            # 逐个用 requests 请求详情页并提取数据
            for idx, detail_url in enumerate(detail_urls):
                try:
                    print(f"  [{idx + 1}/{len(detail_urls)}] 正在抓取: {detail_url}")
                    data = extract_detail(detail_url)
                    batch_data.append(data)
                    counter += 1

                    if counter % 100 == 0:
                        print(f"  [进展] 已抓取 {counter} 条数据")

                    # 每 BATCH_SIZE 条保存一次
                    if len(batch_data) >= BATCH_SIZE:
                        file_index += 1
                        print(f"\n[分批保存] 已达 {BATCH_SIZE} 条，保存第 {file_index} 批...")
                        fp = save_to_excel(batch_data, file_index)
                        if fp:
                            saved_files.append(fp)
                            saved_records.append(len(batch_data))
                        batch_data = []

                    # 请求间隔，防止触发 429
                    time.sleep(REQUEST_DELAY)

                except Exception as e:
                    print(f"  [警告] 抓取失败 ({detail_url}): {e}")
                    # 出错后额外等待，避免连续触发限流
                    time.sleep(REQUEST_DELAY * 2)
                    continue

            # 尝试翻页
            # 先检查是否达到最大页数
            if end_page is not None and page_num >= end_page:
                print(f"  已达测试页数限制 (从第 {start_page} 页开始，共 {max_pages} 页，结束于第 {end_page} 页)，爬取结束。")
                break
            
            try:
                # 翻页前等待，避免请求过快触发限流
                page.wait_for_timeout(5000)
                next_button = page.query_selector(NEXT_PAGE_SELECTOR)
                if not next_button:
                    print("  未找到下一页按钮，爬取结束。")
                    break

                # 检查按钮是否可点击（非禁用状态）
                is_disabled = next_button.evaluate("""el => {
                    const a = el.querySelector('a');
                    if (!a) return true;
                    if (a.getAttribute('aria-disabled') === 'true') return true;
                    if (el.classList.contains('disabled')) return true;
                    return false;
                }""")

                if is_disabled:
                    print("  下一页按钮已禁用，爬取结束。")
                    break

                # 获取下一页链接并直接跳转（避免 AJAX 翻页导致 expect_navigation 超时）
                next_link = next_button.query_selector("a")
                if next_link:
                    next_href = next_link.get_attribute("href")
                    if next_href and not next_href.startswith("http"):
                        next_href = BASE_URL + next_href
                else:
                    next_href = None

                if not next_href:
                    print("  未找到下一页链接，爬取结束。")
                    break

                print(f"  跳转到下一页: {next_href}")

                # 翻页重试计数
                nav_retry_count = 0
                nav_max_retries = 3
                page_loaded = False

                while nav_retry_count <= nav_max_retries:
                    try:
                        # 直接跳转到下一页URL
                        page.goto(next_href, wait_until="domcontentloaded", timeout=120000)

                        # 等待搜索结果出现（比等容器更可靠）
                        try:
                            page.wait_for_selector(SEARCH_RESULT_SELECTOR, timeout=60000)
                            print("  页面加载完成")
                            page_loaded = True
                            break
                        except Exception:
                            # 搜索结果未出现，尝试等容器
                            try:
                                page.wait_for_selector(LIST_SELECTOR, timeout=30000)
                                # 容器出现了但可能内容为空，再等等
                                page.wait_for_timeout(5000)
                                # 再次检查搜索结果
                                result_count = page.evaluate("""() => {
                                    return document.querySelectorAll('#dashboard_notices .search-result').length;
                                }""")
                                if result_count > 0:
                                    print("  页面加载完成")
                                    page_loaded = True
                                    break
                            except Exception:
                                pass

                            if nav_retry_count < nav_max_retries:
                                print(f"  [警告] 页面加载超时，第 {nav_retry_count + 1}/{nav_max_retries} 次重试...")
                                page.wait_for_timeout(5000)
                                nav_retry_count += 1
                            else:
                                print(f"  [错误] 页面加载失败，已重试 {nav_max_retries} 次")
                                break

                    except Exception as e:
                        if nav_retry_count < nav_max_retries:
                            print(f"  [警告] 翻页异常，第 {nav_retry_count + 1}/{nav_max_retries} 次重试... {e}")
                            page.wait_for_timeout(5000)
                            nav_retry_count += 1
                        else:
                            print(f"  [错误] 翻页失败，已重试 {nav_max_retries} 次: {e}")
                            break

                if not page_loaded:
                    print(f"  [提示] 第 {page_num + 1} 页翻页失败，爬取结束")
                    break

                page_num += 1

                # 检查是否达到最大页数
                if end_page is not None and page_num > end_page:
                    print(f"  已达测试页数限制 (从第 {start_page} 页开始，共 {max_pages} 页，结束于第 {end_page} 页)，爬取结束。")
                    break

            except Exception as e:
                print(f"  翻页失败或已到最后一页: {e}")
                break

        # 保存剩余数据
        if batch_data:
            file_index += 1
            print(f"\n[最终保存] 保存剩余 {len(batch_data)} 条数据...")
            fp = save_to_excel(batch_data, file_index)
            if fp:
                saved_files.append(fp)
                saved_records.append(len(batch_data))

        browser.close()

    # 汇总
    print("\n" + "=" * 60)
    print(f"爬取完成！共抓取 {counter} 条数据，保存为 {len(saved_files)} 个文件")
    print(f"保存目录: {SAVE_DIR}")
    print("=" * 60)

    # ====== 判断发送方式 ======

    # 获取所有下载的文件及其大小
    files = []
    for fp in saved_files:
        if os.path.isfile(fp):
            filesize = os.path.getsize(fp)
            files.append((fp, os.path.basename(fp), filesize))

    # 检查是否符合QQ邮箱附件规则
    can_send_as_attachment = True
    total_size = sum(f[2] for f in files)

    # 检查单个文件大小
    for filepath, filename, filesize in files:
        if filesize > QQ_MAX_SINGLE_FILE_SIZE:
            can_send_as_attachment = False
            print(f"\n  ⚠️ 文件 {filename} 大小 {filesize/1024/1024:.1f}MB 超过20MB限制")
            break

    # 检查总大小
    if can_send_as_attachment and total_size > QQ_MAX_TOTAL_SIZE:
        can_send_as_attachment = False
        print(f"\n  ⚠️ 文件总大小 {total_size/1024/1024:.1f}MB 超过50MB限制")

    # 上传并发送邮件
    if sender and auth_code and receiver and saved_files:
        if can_send_as_attachment:
            # 方式1：作为邮件附件发送
            print("\n[步骤] 发送邮件（带附件）...")
            print(f"  文件总大小: {total_size/1024/1024:.1f}MB，符合QQ邮箱附件规则")

            try:
                send_email_with_attachments(
                    subject="英国 Contracts Finder 招标数据",
                    sender=sender,
                    auth_code=auth_code,
                    recipient=receiver,
                    files=files,
                )
            except Exception as e:
                print(f"  邮件发送失败: {e}")

            # 汇总
            print(f"\n{'='*60}")
            print(f"完成! 共抓取 {counter} 条数据，保存为 {len(files)} 个文件")
            print(f"文件已作为附件发送")
            print(f"保存目录: {SAVE_DIR}")

        else:
            # 方式2：上传到Gofile，邮件发送下载链接
            print("\n[步骤] 上传文件到 Gofile.io...")
            download_link, upload_results = upload_to_gofile(saved_files)

            if download_link and upload_results:
                try:
                    send_email_with_gofile_link(
                        subject="英国 Contracts Finder 招标数据 Gofile 下载链接",
                        sender=sender,
                        auth_code=auth_code,
                        recipient=receiver,
                        download_link=download_link,
                        upload_results=upload_results,
                    )
                except Exception as e:
                    print(f"  邮件发送失败: {e}")

                # 汇总
                print(f"\n{'='*60}")
                print(f"完成! 共抓取 {counter} 条数据，保存为 {len(files)} 个文件，上传 {len(upload_results)} 个到Gofile")
                print(f"Gofile 链接: {download_link}")
                print(f"保存目录: {SAVE_DIR}")
            else:
                print("\n[警告] Gofile 上传失败，跳过邮件发送")
    elif saved_files:
        print("\n[提示] 未提供邮件参数，跳过上传和邮件发送")
        print("  如需上传并发送邮件，请使用:")
        print("  python england_tender.py --sender 发件邮箱 --auth-code 授权码 --receiver 收件邮箱")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="英国 Contracts Finder 招标数据爬取、上传与邮件通知"
    )
    parser.add_argument(
        "--sender", required=False, default=None,
        help="发件邮箱（QQ 邮箱）"
    )
    parser.add_argument(
        "--auth-code", required=False, default=None,
        help="发件邮箱 SMTP 授权码"
    )
    parser.add_argument(
        "--receiver", required=False, default=None,
        help="收件邮箱"
    )
    parser.add_argument(
        "--max-pages", required=False, default="2",
        help="最大爬取页数，默认 2 页。可以是数字 n（爬取前 n 页）或 'full'（爬取全部页）"
    )
    parser.add_argument(
        "--start-page", required=False, default="1",
        help="起始页码，默认 1。从指定页码开始爬取"
    )
    args = parser.parse_args()

    # 解析 max_pages 参数
    max_pages_arg = args.max_pages
    if max_pages_arg.lower() == "full":
        max_pages = "full"
    else:
        try:
            max_pages = int(max_pages_arg)
        except ValueError:
            print(f"警告: --max-pages 参数无效 '{args.max_pages}'，使用默认值 2")
            max_pages = 2

    # 解析 start_page 参数
    start_page_arg = args.start_page
    try:
        start_page = int(start_page_arg)
        if start_page < 1:
            print(f"警告: --start-page 参数无效 '{args.start_page}'，使用默认值 1")
            start_page = 1
    except ValueError:
        print(f"警告: --start-page 参数无效 '{args.start_page}'，使用默认值 1")
        start_page = 1

    main(sender=args.sender, auth_code=args.auth_code, receiver=args.receiver, max_pages=max_pages, start_page=start_page)
