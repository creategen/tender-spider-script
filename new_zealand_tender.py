#!/usr/bin/env python3
"""
新西兰 GETS 政府招标网站一键爬取脚本
============================================
功能：爬取4个招标列表页面，逐条提取详情数据，输出带层级合并表头的Excel，
     智能发送邮件（小文件直接附件，大文件上传Gofile.io并发送下载链接）。

用法：
  python3 new_zealand_tender.py --sender 发件邮箱 --auth-code 授权码 --receiver 收件邮箱 [选项]
参数说明：
  --sender        发件邮箱（QQ邮箱）
  --auth-code     发件邮箱SMTP授权码
  --receiver      收件邮箱，多个邮箱用分号分隔（如：xxx@qq.com;yyy@live.com）
选项：
  --full        全量爬取（默认仅爬取"当前招标"）
  --progress    按照进度文件记录的进度爬取（不传则强制重新爬取，忽略进度文件）
  --skip-upload  跳过Gofile上传和邮件发送
  --attachments  传此参数时，文件符合规则则作为邮件附件发送；不传则强制走 Gofile 链接方式

QQ邮箱附件规则：
  - 单个附件 ≤ 20MB
  - 全部附件总大小 ≤ 50MB
  传 --attachments 时，符合规则则直接作为邮件附件发送，否则上传到Gofile.io并发送下载链接。
  不传 --attachments 时，强制上传到Gofile.io并发送下载链接。

依赖：
  pip3 install playwright requests beautifulsoup4 openpyxl
  playwright install chromium
"""

import argparse
import json
import os
import re
import smtplib
import sys
import time
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime, timezone, timedelta
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from playwright.sync_api import sync_playwright

# ==================== 配置 ====================
OUTPUT_DIR = os.path.join(os.getcwd(), "新西兰招标")
PROGRESS_DIR = os.path.join(os.getcwd(), "新西兰招标/.progress")

SITES = [
    {"url": "https://www.gets.govt.nz/ExternalIndex.htm", "category": "当前招标"},
    {"url": "https://www.gets.govt.nz/ExternalLateTenderList.htm", "category": "逾期招标"},
    {"url": "https://www.gets.govt.nz/ExternalClosedTenderList.htm", "category": "已关闭招标"},
    {"url": "https://www.gets.govt.nz/ExternalAwardedTenderList.htm", "category": "已完成招标"},
]

# QQ邮箱附件限制
QQ_MAX_SINGLE_FILE_SIZE = 20 * 1024 * 1024  # 20MB
QQ_MAX_TOTAL_SIZE = 50 * 1024 * 1024  # 50MB

# 性能参数
DETAIL_DELAY = 0.15       # HTTP详情请求间隔(秒)
PAGE_DELAY = 0.3          # Playwright翻页间隔(秒)
BATCH_SIZE = 50           # 每批处理条数
MAX_WORKERS = 5           # 并发提取线程数

# HTTP请求Session
HTTP_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.5",
}
http_session = requests.Session()
http_session.headers.update(HTTP_HEADERS)


# ==================== 工具函数 ====================

def clean_string(s):
    """清理XML非法控制字符，避免openpyxl报错"""
    if not isinstance(s, str):
        return s
    return re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]', '', s)


def load_progress(progress_file):
    """加载进度"""
    if os.path.exists(progress_file):
        with open(progress_file, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def save_progress(progress, progress_file):
    """保存进度"""
    os.makedirs(os.path.dirname(progress_file), exist_ok=True)
    with open(progress_file, "w", encoding="utf-8") as f:
        json.dump(progress, f, ensure_ascii=False)


# ==================== 第一步：Playwright收集链接 ====================

def get_all_tender_links(base_url):
    """使用Playwright翻页收集列表页所有招标链接"""
    all_links = []
    with sync_playwright() as p:
        browser = p.chromium.launch(
            headless=True,
            args=['--no-sandbox', '--disable-setuid-sandbox', '--disable-dev-shm-usage']
        )
        page = browser.new_page()
        page.goto(base_url, timeout=60000)
        page.wait_for_selector("#treetbl", timeout=30000)

        page_num = 1
        while True:
            page.wait_for_selector("#treetbl", timeout=30000)
            page.wait_for_timeout(300)

            links = page.evaluate("""() => {
                const rows = document.querySelectorAll('tr.tender[id^="tender-"]');
                return Array.from(rows).map(row => {
                    const a = row.querySelector('a');
                    return a ? a.href : null;
                }).filter(x => x);
            }""")

            if not links:
                break
            all_links.extend(links)

            if page_num % 50 == 0:
                print(f"    已扫描 {page_num} 页, {len(all_links)} 条链接")

            next_btn = page.query_selector("#next-active")
            if not next_btn:
                break
            next_btn.click()
            page.wait_for_timeout(int(PAGE_DELAY * 1000))
            page_num += 1

        browser.close()

    # 去重
    seen = set()
    unique = []
    for link in all_links:
        if link not in seen:
            seen.add(link)
            unique.append(link)

    print(f"    总计获取 {len(unique)} 条唯一招标链接")
    return unique


# ==================== 第二步：HTTP提取详情 ====================

def extract_detail_data(url):
    """使用HTTP请求提取详情页数据（速度快）"""
    try:
        resp = http_session.get(url, timeout=30)
        if resp.status_code != 200:
            return None

        soup = BeautifulSoup(resp.text, 'html.parser')

        # 提取 Details 表格
        table = soup.select_one('#yui-main > div > div > div > table')
        if not table:
            return None

        details = {}
        for row in table.find_all('tr'):
            cells = row.find_all('td')
            if len(cells) >= 2:
                key = cells[0].get_text(strip=True).replace('\xa0', '').rstrip(':').strip()
                value = cells[1].get_text(strip=True)
                if key:
                    details[key] = value

        if not details:
            return None

        # 提取 Overview
        overview = ""
        divider = soup.select_one('.detail-divider')
        if divider:
            text_parts = []
            sibling = divider.find_next_sibling()
            while sibling and 'detail-divider' not in (sibling.get('class') or []):
                t = sibling.get_text(strip=True)
                if t:
                    text_parts.append(t)
                sibling = sibling.find_next_sibling()
            overview = '\n'.join(text_parts)

        return {"URL": url, "Details": details, "Overview": overview}
    except Exception:
        return None


def extract_details_batch(urls):
    """并发批量提取详情"""
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        results = list(executor.map(extract_detail_data, urls))
    return results


# ==================== 第三步：生成Excel ====================

def create_excel(data_list, category, output_path):
    """创建带层级合并表头的Excel文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = category[:31]

    if not data_list:
        ws.cell(row=1, column=1, value="URL")
        ws.cell(row=1, column=2, value="Details")
        ws.cell(row=1, column=3, value="Overview")
        wb.save(output_path)
        return

    # 收集所有二级表头
    all_sub_fields = set()
    for item in data_list:
        if item and item.get("Details"):
            all_sub_fields.update(item["Details"].keys())

    sub_fields = sorted(all_sub_fields)
    num_sub_fields = len(sub_fields)
    overview_col = 2 + num_sub_fields

    # 样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    sub_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
    sub_font = Font(bold=True, color="FFFFFF", size=10)
    border = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
    wrap = Alignment(wrap_text=True, vertical='top')
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # 第一行：一级表头
    for col, val in [(1, "URL"), (2, "Details"), (overview_col, "Overview")]:
        c = ws.cell(row=1, column=col, value=val)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = border

    if num_sub_fields > 1:
        ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=1 + num_sub_fields)
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    ws.merge_cells(start_row=1, start_column=overview_col, end_row=2, end_column=overview_col)

    for col in range(2, 2 + num_sub_fields):
        c = ws.cell(row=1, column=col)
        c.fill = header_fill
        c.border = border

    # 第二行：二级表头
    for i, field in enumerate(sub_fields):
        c = ws.cell(row=2, column=2 + i, value=field)
        c.font = sub_font
        c.fill = sub_fill
        c.alignment = center
        c.border = border

    ws.cell(row=2, column=1).border = border
    ws.cell(row=2, column=1).fill = header_fill
    ws.cell(row=2, column=overview_col).border = border
    ws.cell(row=2, column=overview_col).fill = header_fill

    # 数据行
    for row_idx, item in enumerate(data_list, start=3):
        if not item:
            continue
        c = ws.cell(row=row_idx, column=1, value=clean_string(item.get("URL", "")))
        c.alignment = wrap
        c.border = border

        details = item.get("Details", {})
        for i, field in enumerate(sub_fields):
            c = ws.cell(row=row_idx, column=2 + i, value=clean_string(details.get(field, "")))
            c.alignment = wrap
            c.border = border

        c = ws.cell(row=row_idx, column=overview_col, value=clean_string(item.get("Overview", "")))
        c.alignment = wrap
        c.border = border

    # 列宽
    for col in range(1, overview_col + 1):
        max_len = 0
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=False):
            for cell in row:
                if cell.value:
                    for line in str(cell.value).split('\n'):
                        max_len = max(max_len, sum(2 if ord(c) > 127 else 1 for c in line))
        ws.column_dimensions[get_column_letter(col)].width = min(max(max_len + 2, 10), 50)

    ws.freeze_panes = "A3"
    wb.save(output_path)
    print(f"    Excel 已保存: {output_path} (共 {len(data_list)} 条记录)")


# ==================== 第四步：爬取主流程 ====================

def scrape_category(site_config, progress_file, use_progress=False):
    """爬取单个类别的完整流程"""
    url = site_config["url"]
    category = site_config["category"]
    progress_key = category

    print(f"\n{'=' * 60}")
    print(f"开始爬取: {category} - {url}")
    print(f"{'=' * 60}")

    # 加载进度文件
    progress = load_progress(progress_file)
    
    # 如果不使用进度，忽略进度文件中的 completed 状态，强制重新爬取
    if not use_progress:
        saved = {"all_urls": [], "data_list": [], "completed": False}
    else:
        saved = progress.get(progress_key, {"all_urls": [], "data_list": [], "completed": False})

    if saved.get("completed") and os.path.exists(saved.get("output_path", "")):
        print(f"  {category} 已完成，跳过")
        return saved.get("output_path")

    all_urls = saved.get("all_urls", [])
    data_list = saved.get("data_list", [])
    done_urls = {item["URL"] for item in data_list if item}

    # 收集链接（如果没有的话）
    if not all_urls:
        print(f"  正在收集 {category} 的所有链接...")
        all_urls = get_all_tender_links(url)
    else:
        print(f"  恢复已保存的链接: {len(all_urls)} 条")

    # 提取缺失的详情
    pending = [u for u in all_urls if u not in done_urls]
    print(f"  总计 {len(all_urls)} 条, 已提取 {len(data_list)} 条, 待提取 {len(pending)} 条")

    if pending:
        total_batches = (len(pending) + BATCH_SIZE - 1) // BATCH_SIZE
        for batch_start in range(0, len(pending), BATCH_SIZE):
            batch = pending[batch_start:batch_start + BATCH_SIZE]
            batch_num = batch_start // BATCH_SIZE + 1

            results = extract_details_batch(batch)

            success = 0
            for link, result in zip(batch, results):
                if result:
                    data_list.append(result)
                    success += 1

            if batch_num % 5 == 0 or batch_num == total_batches:
                print(f"  批次 {batch_num}/{total_batches}: 成功 {success}/{len(batch)}, "
                      f"累计 {len(data_list)}/{len(all_urls)}")

            # 保存进度
            progress[progress_key] = {
                "all_urls": all_urls,
                "data_list": data_list,
                "completed": False,
            }
            save_progress(progress, progress_file)

            time.sleep(DETAIL_DELAY)

    # 生成Excel
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    # 使用东八区时间
    tz_beijing = timezone(timedelta(hours=8))
    now = datetime.now(tz_beijing)
    timestamp = now.strftime("%Y-%m-%d-%H-%M-%S")
    filename = f"新西兰-{category}-{timestamp}.xlsx"
    output_path = os.path.join(OUTPUT_DIR, filename)
    create_excel(data_list, category, output_path)

    # 标记完成
    progress[progress_key] = {
        "all_urls": all_urls,
        "data_list": data_list,
        "completed": True,
        "output_path": output_path,
    }
    save_progress(progress, progress_file)

    return output_path


# ==================== 第五步：上传Gofile ====================

def upload_to_gofile(file_paths):
    """将多个文件上传到Gofile同一目录，返回下载链接"""
    print("\n上传文件到 Gofile.io...")

    # 创建账户
    create_resp = requests.post("https://api.gofile.io/accounts", timeout=30)
    create_data = create_resp.json()
    if create_data.get("status") != "ok":
        print(f"  创建Gofile账户失败: {create_data}")
        return None

    token = create_data["data"]["token"]
    root_folder = create_data["data"]["rootFolder"]

    # 获取服务器
    server_resp = requests.get("https://api.gofile.io/servers", timeout=30)
    server = server_resp.json()["data"]["servers"][0]["name"]

    folder_id = None
    folder_code = None

    for i, fpath in enumerate(file_paths):
        fname = os.path.basename(fpath)
        size_mb = os.path.getsize(fpath) / 1024 / 1024
        print(f"  上传: {fname} ({size_mb:.1f}MB)...")

        data = {"token": token, "folderId": root_folder if i == 0 else folder_id}

        # 每次获取新服务器（避免切换问题）
        sr = requests.get("https://api.gofile.io/servers", timeout=30)
        s = sr.json()["data"]["servers"][0]["name"]

        with open(fpath, "rb") as f:
            upload_resp = requests.post(
                f"https://{s}.gofile.io/uploadFile",
                files={"file": (fname, f)},
                data=data,
                timeout=300,
            )

        upload_data = upload_resp.json()
        if upload_data.get("status") == "ok":
            if i == 0:
                folder_id = upload_data["data"]["parentFolder"]
                folder_code = upload_data["data"]["parentFolderCode"]
            print(f"    ✓ 成功")
        else:
            print(f"    ✗ 失败: {upload_data}")

        time.sleep(1)

    # 设置文件夹公开访问（必须！）
    if folder_id and token:
        print("  设置文件夹公开访问...")
        pub_resp = requests.put(
            f"https://api.gofile.io/contents/{folder_id}/update",
            headers={"Authorization": f"Bearer {token}"},
            json={"attribute": "public", "attributeValue": True},
            timeout=30,
        )
        if pub_resp.status_code == 200 and pub_resp.json().get("status") == "ok":
            print("    ✓ 文件夹已设为公开")
        else:
            print(f"    ✗ 设置公开失败: {pub_resp.text[:200]}")

    download_link = f"https://gofile.io/d/{folder_code}" if folder_code else None
    print(f"  下载链接: {download_link}")
    return download_link


# ==================== 第六步：发送邮件 ====================

def send_email_with_gofile_link(sender, auth_code, receiver, download_link, file_info_list):
    """
    发送包含Gofile下载链接的邮件（HTML格式）
    receiver: 单个邮箱字符串，或多个邮箱用分号分隔的字符串
    """
    # 解析多个收件人（用分号分隔）
    if ';' in receiver:
        receiver_list = [r.strip() for r in receiver.split(';') if r.strip()]
    else:
        receiver_list = [receiver.strip()]
    
    print(f"\n发送邮件到 {len(receiver_list)} 个收件人: {', '.join(receiver_list)}...")

    subject = "新西兰GETS政府招标网站爬取结果 - Excel下载链接"

    # 文件清单表格行
    file_rows_html = ""
    file_rows_text = ""
    for i, info in enumerate(file_info_list, 1):
        bg = "#f2f2f2" if i % 2 == 1 else "#ffffff"
        file_rows_html += f"""<tr style="background-color: {bg};">
<td style="border: 1px solid #ddd; padding: 10px;">{i}</td>
<td style="border: 1px solid #ddd; padding: 10px;">{info['filename']}</td>
<td style="border: 1px solid #ddd; padding: 10px;">{info['category']}</td>
<td style="border: 1px solid #ddd; padding: 10px;">{info['records']:,}</td></tr>
"""
        file_rows_text += f"{i}. {info['filename']} ({info['category']}, {info['records']}条)\n"

    now_str = datetime.now(timezone(timedelta(hours=8))).strftime('%Y年%m月%d日')

    html_body = f"""<html><body style="font-family: Microsoft YaHei, Arial, sans-serif; color: #333; font-size: 14px; line-height: 1.8; max-width: 700px; margin: 0 auto;">
<h2 style="color: #1a5276; border-bottom: 2px solid #2980b9; padding-bottom: 8px; font-size: 20px;">新西兰 GETS 政府招标网站爬取结果</h2>
<div style="background-color: #eaf2f8; padding: 15px; border-radius: 5px; margin: 15px 0; font-size: 14px;">
<strong>爬取时间：</strong>{now_str}<br>
<strong>数据来源：</strong>https://www.gets.govt.nz<br>
<strong>总计：</strong>{len(file_info_list)}个Excel文件，{sum(f['records'] for f in file_info_list):,}条招标记录
</div>
<h3 style="font-size: 16px;">下载链接</h3>
<div style="background-color: #fff3cd; border: 1px solid #ffc107; padding: 20px; border-radius: 5px; text-align: center; margin: 15px 0;">
<a href="{download_link}" style="font-size: 18px; color: #2980b9; font-weight: bold; text-decoration: none;">点击下载全部文件</a>
<br><span style="color: #888; font-size: 13px;">{download_link}</span>
<br><span style="color: #666; font-size: 12px;">（{len(file_info_list)}个Excel文件均在同一目录下，可逐个下载）</span>
</div>
<h3 style="font-size: 16px;">文件清单</h3>
<table style="border-collapse: collapse; width: 100%; font-size: 14px;">
<tr style="background-color: #2980b9; color: white;"><th style="padding: 10px; text-align: left;">序号</th><th style="padding: 10px; text-align: left;">文件名</th><th style="padding: 10px; text-align: left;">招标类别</th><th style="padding: 10px; text-align: left;">记录数</th></tr>
{file_rows_html}
</table>
<div style="background-color: #f8d7da; border: 1px solid #f5c6cb; padding: 15px; border-radius: 5px; margin: 20px 0; font-size: 14px;">
<strong style="color: #721c24;">提醒：Gofile.io 链接有效期有限，建议尽快下载。</strong>
</div>
<p style="color:#888; font-size:12px; margin-top:20px;">此邮件由自动化爬取系统发送，请勿直接回复。</p>
</body></html>"""

    text_body = f"""新西兰 GETS 政府招标网站爬取结果

爬取时间：{now_str}
数据来源：https://www.gets.govt.nz
总计：{len(file_info_list)}个Excel文件，{sum(f['records'] for f in file_info_list):,}条招标记录

下载链接（{len(file_info_list)}个Excel文件在同一目录下，可逐个下载）：
{download_link}

文件清单：
{file_rows_text}
提醒：Gofile.io 链接有效期有限，建议尽快下载。
"""

    msg = MIMEMultipart("alternative")
    msg["From"] = sender
    msg["To"] = ", ".join(receiver_list)  # 多个收件人用逗号分隔
    msg["Subject"] = subject

    msg.attach(MIMEText(text_body, "plain", "utf-8"))
    msg.attach(MIMEText(html_body, "html", "utf-8"))

    try:
        server = smtplib.SMTP_SSL("smtp.qq.com", 465)
        server.login(sender, auth_code)
        server.sendmail(sender, receiver_list, msg.as_string())
        server.quit()
        print("  ✓ 邮件发送成功！")
        return True
    except Exception as e:
        # 尝试TLS方式
        try:
            server = smtplib.SMTP("smtp.qq.com", 587)
            server.starttls()
            server.login(sender, auth_code)
            server.sendmail(sender, receiver_list, msg.as_string())
            server.quit()
            print("  ✓ 邮件发送成功(TLS)！")
            return True
        except Exception as e2:
            print(f"  ✗ 邮件发送失败: {e2}")
            return False


def send_email_with_attachments(sender, auth_code, receiver, file_info_list, output_files):
    """
    发送带附件的邮件（符合QQ邮箱规则）
    receiver: 单个邮箱字符串，或多个邮箱用分号分隔的字符串
    """
    # 解析多个收件人（用分号分隔）
    if ';' in receiver:
        receiver_list = [r.strip() for r in receiver.split(';') if r.strip()]
    else:
        receiver_list = [receiver.strip()]
    
    print(f"\n发送邮件到 {len(receiver_list)} 个收件人: {', '.join(receiver_list)}（带附件）...")

    subject = "新西兰GETS政府招标网站爬取结果 - Excel附件"

    # 文件清单
    file_rows_text = ""
    for i, info in enumerate(file_info_list, 1):
        file_rows_text += f"{i}. {info['filename']} ({info['category']}, {info['records']}条)\n"

    now_str = datetime.now(timezone(timedelta(hours=8))).strftime('%Y年%m月%d日')

    html_body = f"""<html><body style="font-family: Microsoft YaHei, Arial, sans-serif; color: #333; font-size: 14px; line-height: 1.8; max-width: 700px; margin: 0 auto;">
<h2 style="color: #1a5276; border-bottom: 2px solid #2980b9; padding-bottom: 8px; font-size: 20px;">新西兰 GETS 政府招标网站爬取结果</h2>
<div style="background-color: #eaf2f8; padding: 15px; border-radius: 5px; margin: 15px 0; font-size: 14px;">
<strong>爬取时间：</strong>{now_str}<br>
<strong>数据来源：</strong>https://www.gets.govt.nz<br>
<strong>总计：</strong>{len(file_info_list)}个Excel文件，{sum(f['records'] for f in file_info_list):,}条招标记录（作为附件发送）
</div>
<h3 style="font-size: 16px;">文件清单</h3>
<pre style="background-color: #f5f5f5; padding: 15px; border-radius: 5px; font-family: Consolas, Monaco, monospace; font-size: 13px;">{file_rows_text}</pre>
<div style="background-color: #d4edda; border: 1px solid #c3e6cb; padding: 15px; border-radius: 5px; margin: 20px 0; font-size: 14px;">
<strong style="color: #155724;">✓ 文件已作为附件发送，可直接下载。</strong>
</div>
<p style="color:#888; font-size:12px; margin-top:20px;">此邮件由自动化爬取系统发送，请勿直接回复。</p>
</body></html>"""

    text_body = f"""新西兰 GETS 政府招标网站爬取结果

爬取时间：{now_str}
数据来源：https://www.gets.govt.nz
总计：{len(file_info_list)}个Excel文件，{sum(f['records'] for f in file_info_list):,}条招标记录（作为附件发送）

文件清单：
{file_rows_text}
文件已作为附件发送，可直接下载。
"""

    msg = MIMEMultipart("mixed")
    msg["From"] = sender
    msg["To"] = ", ".join(receiver_list)  # 多个收件人用逗号分隔
    msg["Subject"] = subject

    msg.attach(MIMEText(text_body, "plain", "utf-8"))
    msg.attach(MIMEText(html_body, "html", "utf-8"))

    # 添加附件
    for i, fpath in enumerate(output_files):
        filename = os.path.basename(fpath)
        try:
            with open(fpath, "rb") as f:
                part = MIMEBase("application", "octet-stream")
                part.set_payload(f.read())
                encoders.encode_base64(part)
                # 正确设置附件文件名
                part.add_header('Content-Disposition', 'attachment', filename=('utf-8', '', filename))
                msg.attach(part)
            print(f"  已添加附件: {filename}")
        except Exception as e:
            print(f"  添加附件失败: {filename} - {e}")

    try:
        server = smtplib.SMTP_SSL("smtp.qq.com", 465)
        server.login(sender, auth_code)
        server.sendmail(sender, receiver_list, msg.as_string())
        server.quit()
        print("  ✓ 邮件发送成功（带附件）！")
        return True
    except Exception as e:
        # 尝试TLS方式
        try:
            server = smtplib.SMTP("smtp.qq.com", 587)
            server.starttls()
            server.login(sender, auth_code)
            server.sendmail(sender, receiver_list, msg.as_string())
            server.quit()
            print("  ✓ 邮件发送成功（带附件，TLS）！")
            return True
        except Exception as e2:
            print(f"  ✗ 邮件发送失败: {e2}")
            return False


# ==================== 主入口 ====================

def main():
    parser = argparse.ArgumentParser(description="新西兰GETS政府招标网站爬取工具")
    parser.add_argument("--sender", required=True, help="发件邮箱（QQ邮箱）")
    parser.add_argument("--auth-code", required=True, help="发件邮箱SMTP授权码")
    parser.add_argument("--receiver", required=True, help="收件邮箱")
    parser.add_argument("--output-dir", default=None, help="输出目录（默认: 脚本所在目录/output）")
    parser.add_argument("--skip-upload", action="store_true", help="跳过Gofile上传和邮件发送")
    parser.add_argument("--full", action="store_true", help="全量爬取（默认仅爬取当前招标）")
    parser.add_argument("--progress", action="store_true", help="按照进度文件记录的进度爬取（不传则强制重新爬取，忽略进度文件）")
    parser.add_argument(
        "--attachments", action="store_true",
        help="传此参数时，文件符合规则则作为邮件附件发送；不传则强制走 Gofile 链接方式"
    )
    args = parser.parse_args()

    # 设置输出目录
    output_dir = args.output_dir if args.output_dir else OUTPUT_DIR

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    os.makedirs(PROGRESS_DIR, exist_ok=True)

    progress_file = os.path.join(PROGRESS_DIR, "gets_progress.json")

    print("=" * 60)
    print("新西兰 GETS 政府招标网站爬取工具")
    print(f"开始时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"输出目录: {OUTPUT_DIR}")
    print("=" * 60)

    # ===== 爬取类别 =====
    sites_to_scrape = SITES if args.full else [SITES[0]]
    if args.full:
        print("爬取模式: 全量爬取（4个类别）")
    else:
        print("爬取模式: 仅爬取当前招标（传 --full 可全量爬取）")

    output_files = []
    for site in sites_to_scrape:
        output_path = scrape_category(site, progress_file, args.progress)
        if output_path and os.path.exists(output_path):
            output_files.append(output_path)

    print(f"\n{'=' * 60}")
    print(f"爬取完成！生成 {len(output_files)} 个文件:")
    for f in output_files:
        print(f"  - {f}")

    if not output_files:
        print("没有生成任何文件，退出")
        return

    if args.skip_upload:
        print("\n跳过Gofile上传和邮件发送（--skip-upload）")
        return

    # ===== 统计文件信息和大小 =====
    file_info_list = []
    output_files_with_size = []  # (filepath, filename, filesize)
    total_size = 0
    can_send_as_attachment = True

    for fpath in output_files:
        fname = os.path.basename(fpath)
        filesize = os.path.getsize(fpath)
        total_size += filesize
        output_files_with_size.append((fpath, fname, filesize))

        # 从文件名提取类别
        category = "未知"
        for site in sites_to_scrape:
            if site["category"] in fname:
                category = site["category"]
                break

        # 统计记录数
        from openpyxl import load_workbook
        try:
            wb = load_workbook(fpath, read_only=True)
            ws = wb.active
            records = max(ws.max_row - 2, 0)
            wb.close()
        except Exception:
            records = 0

        file_info_list.append({
            "filename": fname,
            "category": category,
            "records": records,
        })

    # ===== 判断发送方式 =====
    # 判断发送方式
    # - 不传 --attachments：强制走 Gofile 方式
    # - 传 --attachments：按现有规则判断
    if not args.attachments:
        # 强制走 Gofile 方式
        can_send_as_attachment = False
    else:
        # 检查是否符合QQ邮箱附件规则
        can_send_as_attachment = True

        # 检查单个文件大小
        for fpath, fname, filesize in output_files_with_size:
            if filesize > QQ_MAX_SINGLE_FILE_SIZE:
                can_send_as_attachment = False
                print(f"\n  ⚠️ 文件 {fname} 大小 {filesize/1024/1024:.1f}MB 超过20MB限制")
                break

        # 检查总大小
        if can_send_as_attachment and total_size > QQ_MAX_TOTAL_SIZE:
            can_send_as_attachment = False
            print(f"\n  ⚠️ 文件总大小 {total_size/1024/1024:.1f}MB 超过50MB限制")

    if can_send_as_attachment:
        # 方式1：作为邮件附件发送
        print(f"\n文件总大小: {total_size/1024/1024:.1f}MB，符合QQ邮箱附件规则")
        send_email_with_attachments(args.sender, args.auth_code, args.receiver, file_info_list, output_files)

        print(f"\n{'=' * 60}")
        print("全部完成！")
        print(f"结束时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"文件已作为附件发送")
        print(f"{'=' * 60}")

    else:
        # 方式2：上传到Gofile，邮件发送下载链接
        download_link = upload_to_gofile(output_files)

        if not download_link:
            print("上传失败，跳过邮件发送")
            return

        send_email_with_gofile_link(args.sender, args.auth_code, args.receiver, download_link, file_info_list)

        print(f"\n{'=' * 60}")
        print("全部完成！")
        print(f"结束时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"下载链接: {download_link}")
        print(f"{'=' * 60}")


if __name__ == "__main__":
    main()
