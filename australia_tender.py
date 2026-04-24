#!/usr/bin/env python3
"""
澳大利亚 AusTender 招标数据爬取脚本
从 https://www.tenders.gov.au 爬取招标信息

功能：
  - 网站爬取1：从关键词搜索页面爬取 ATM 列表
  - 网站爬取2：从 RSS XML 解析并爬取采购项目
  - 提取详情页所有字段
  - 保存到 Excel
  - 上传至 Gofile.io 并发送下载链接邮件

详情页类型说明：
  三种类型的详情页：
    /Atm/Show/ - ATM (Approach to Market) 市场招标
    /Son/Show/ - SON (Standing Offer Notice) 长期报价通知
    /Cn/Show/ - CN (Contract Notice) 合同通知

用法:
# 默认：执行两种爬取，获取所有状态的数据
python australia_tender.py

# 仅执行网站爬取1（搜索页面）
python australia_tender.py --type 1

# 仅执行网站爬取2（RSS XML）
python australia_tender.py --type 2

# 状态筛选：full=全部, close=仅Closed, open=仅非Closed
python australia_tender.py --search-status close
python australia_tender.py --rss-status open

# 仅爬取指定类型的详情页（不传则爬取全部类型）
python australia_tender.py --tender-type Atm,Son
python australia_tender.py --tender-type Cn

# 指定页数
python australia_tender.py --start-page 1 --max-pages 5

# 带邮件发送
python australia_tender.py --sender xxx@qq.com --auth-code xxx --receiver xxx@qq.com

# 传 --attachments 时，文件符合规则则作为邮件附件发送；不传则强制走 Gofile 链接方式
python australia_tender.py --sender xxx@qq.com --auth-code xxx --receiver xxx@qq.com --attachments

# 完整示例：爬取搜索页面，仅Atm和Son类型，仅Closed状态，从第2页开始爬3页，并发送邮件
python australia_tender.py --type 0 --tender-type Atm,Son,Cn --search-status full --rss-status close --start-page 1 --max-pages full --sender xxx@qq.com --auth-code xxx --receiver xxx@qq.com

参数说明：
  --type          爬取类型：不传或传0执行两种爬取，1仅搜索页面，2仅RSS XML
  --tender-type   详情页类型筛选：不传则爬取全部类型，可指定 Atm/Son/Cn（多个用逗号分隔）
  --search-status 搜索页面状态筛选：full=全部（默认）, close=仅Closed, open=仅非Closed
  --rss-status    RSS爬取状态筛选：full=全部（默认）, close=仅Closed, open=仅非Closed
  --start-page    起始页码，默认1
  --max-pages     最大爬取页数，默认full（全部页）
  --sender        发件邮箱（QQ邮箱）
  --auth-code     发件邮箱SMTP授权码
  --receiver      收件邮箱，多个邮箱用分号分隔（如：xxx@qq.com;yyy@live.com）
  --attachments   传此参数时，文件符合规则则作为邮件附件发送；不传则强制走 Gofile 链接方式

QQ邮箱附件规则：
  - 单个附件 ≤ 20MB
  - 全部附件总大小 ≤ 50MB
  传 --attachments 时，符合规则则直接作为邮件附件发送，否则上传到Gofile.io并发送下载链接。
  不传 --attachments 时，强制上传到Gofile.io并发送下载链接。
"""

import argparse
import os
import smtplib
import xml.etree.ElementTree as ET
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from urllib3.exceptions import InsecureRequestWarning
import urllib3

import requests
from bs4 import BeautifulSoup
from playwright.sync_api import sync_playwright
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# 禁用 SSL 警告
urllib3.disable_warnings(InsecureRequestWarning)

# ==================== 固定配置 ====================

# 网站爬取1 配置
SEARCH_URL = "https://www.tenders.gov.au/Search/KeywordSearch?keyword=ATM2"
BASE_URL = "https://www.tenders.gov.au"
LIST_SELECTOR = "article"  # 每个搜索结果是一个 <article> 元素
NEXT_PAGE_SELECTOR = 'a[aria-label="Next page"]'

# 网站爬取2 配置
RSS_URL = "https://www.tenders.gov.au/public_data/rss/rss.xml"

VIEWPORT = {"width": 1280, "height": 800}
USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/120.0.0.0 Safari/537.36"
)
BATCH_SIZE = 10000
REQUEST_DELAY = 1.5
MAX_RETRIES = 3
RETRY_DELAYS = [121, 121, 0]

# 保存目录
SAVE_DIR = os.path.join(os.getcwd(), "澳大利亚招标")

# QQ邮箱附件限制
QQ_MAX_SINGLE_FILE_SIZE = 20 * 1024 * 1024  # 20MB
QQ_MAX_TOTAL_SIZE = 50 * 1024 * 1024  # 50MB

# ==================== HTTP 会话 ====================

http_session = requests.Session()
http_session.headers.update({"User-Agent": USER_AGENT})
http_session.verify = False

# ==================== 表头定义 ====================

# 网站爬取1 表头（当前采购项目）
HEADER_STRUCTURE_1 = [
    ("Detail URL", []),
    ("Title", []),
    ("Contact Details", []),
    ("Email", []),
    ("ATM ID", []),
    ("Agency", []),
    ("Category", []),
    ("Close Date & Time", []),
    ("Publish Date", []),
    ("Location", []),
    ("ATM Type", []),
    ("Multi Agency Access", []),
    ("Panel Arrangement", []),
    ("Multi-stage", []),
    ("Multi-stage Criteria", []),
    ("Description", []),
    ("Other Instructions", []),
    ("Conditions for Participation", []),
    ("Timeframe for Delivery", []),
    ("Address for Lodgement", []),
    ("Addenda Available", []),
]

# 网站爬取2 表头（RSS采购数据）- 与爬取1相同
HEADER_STRUCTURE_2 = HEADER_STRUCTURE_1.copy()

FLAT_HEADERS_1 = []
for parent, children in HEADER_STRUCTURE_1:
    if children:
        for child in children:
            FLAT_HEADERS_1.append(child)
    else:
        FLAT_HEADERS_1.append(parent)

FLAT_HEADERS_2 = FLAT_HEADERS_1.copy()


# ==================== Excel 写入 ====================

def save_to_excel(batch_data, file_path, flat_headers):
    """
    将数据保存到 Excel 文件
    """
    if not batch_data:
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = "招标数据"

    # 样式定义
    header_font = Font(name="Arial", bold=True, size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    data_alignment = Alignment(vertical="top", wrap_text=True)

    # 写入表头
    for col_idx, header in enumerate(flat_headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font_white
        cell.alignment = header_alignment
        cell.fill = header_fill
        cell.border = thin_border

    # 写入数据
    for row_idx, record in enumerate(batch_data, start=2):
        for col_idx, key in enumerate(flat_headers, start=1):
            value = record.get(key, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = data_alignment
            cell.border = thin_border

    # 自动调整列宽
    for col_idx in range(1, len(flat_headers) + 1):
        max_length = 0
        col_letter = get_column_letter(col_idx)
        for row in ws.iter_rows(min_row=1, max_row=min(len(batch_data) + 1, 100),
                                min_col=col_idx, max_col=col_idx):
            for cell in row:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
        adjusted_width = min(max_length + 2, 60)
        ws.column_dimensions[col_letter].width = max(adjusted_width, 10)

    # 冻结首行
    ws.freeze_panes = "A2"

    # 保存文件
    try:
        wb.save(file_path)
        print(f"\n[保存] 已保存 {len(batch_data)} 条数据到 {file_path}\n")
        return file_path
    except Exception as e:
        print(f"[错误] 保存文件失败: {e}")
        return None


# ==================== 详情页数据提取 ====================

def fetch_detail_page(url):
    """
    用 requests 获取详情页 HTML 并解析为 BeautifulSoup 对象
    """
    for attempt in range(MAX_RETRIES):
        try:
            response = http_session.get(url, timeout=30)
            response.raise_for_status()
            break
        except requests.exceptions.HTTPError as e:
            if response.status_code == 429:
                wait_time = RETRY_DELAYS[attempt]
                print(f"    [429] 请求过快，第 {attempt + 1}/{MAX_RETRIES - 1} 次重试，等待 {wait_time}s...")
                import time
                time.sleep(wait_time)
                if attempt == MAX_RETRIES - 1:
                    raise
            else:
                raise
        except requests.exceptions.ConnectionError:
            wait_time = RETRY_DELAYS[attempt]
            print(f"    [连接错误] 第 {attempt + 1}/{MAX_RETRIES - 1} 次重试，等待 {wait_time}s...")
            import time
            time.sleep(wait_time)
            if attempt == MAX_RETRIES - 1:
                raise

    soup = BeautifulSoup(response.text, "html.parser")
    return soup


def check_closed_status(soup):
    """
    检查 div#mainInnerVisual 是否包含 Closed
    """
    main_visual = soup.find("div", id="mainInnerVisual")
    if main_visual:
        text = main_visual.get_text()
        return "Closed" in text
    return False


def extract_detail(url, flat_headers, status_filter="full"):
    """
    从详情页提取所有目标字段
    
    status_filter: 状态筛选
        "full": 保存所有数据
        "close": 仅保存包含Closed的数据
        "open": 仅保存不包含Closed的数据
    """
    import time
    data = {key: "" for key in flat_headers}
    data["Detail URL"] = url

    try:
        soup = fetch_detail_page(url)

        # 检查是否包含 Closed
        has_closed = check_closed_status(soup)
        
        # 根据状态筛选决定是否跳过
        if status_filter == "close" and not has_closed:
            print(f"    [跳过] 不包含 Closed 状态")
            return None
        elif status_filter == "open" and has_closed:
            print(f"    [跳过] 包含 Closed 状态")
            return None
        
        # 记录状态信息
        status_info = "Closed" if has_closed else "非Closed"
        print(f"    [状态] {status_info}")

        # Title: p标签且role="heading"
        title_elem = soup.find("p", role="heading")
        if title_elem:
            data["Title"] = title_elem.get_text(strip=True)

        # Contact Details 和 Email: div.contact-long
        contact_div = soup.find("div", class_="contact-long")
        if contact_div:
            # 提取 Contact Details
            contact_label = contact_div.find(string=lambda x: x and "Contact Details" in x if x else False)
            if contact_label:
                # 找到 Contact Details 后面的内容
                parent = contact_label.parent if hasattr(contact_label, 'parent') else contact_div
                next_text = parent.get_text(strip=True)
                if ":" in next_text:
                    data["Contact Details"] = next_text.split(":", 1)[1].strip()
            
            # 提取所有文本内容作为 Contact Details
            full_text = contact_div.get_text(separator="\n", strip=True)
            lines = full_text.split("\n")
            for line in lines:
                if "Contact Details" in line and ":" in line:
                    data["Contact Details"] = line.split(":", 1)[1].strip()
                if "@" in line and "." in line:
                    # 简单的邮箱提取
                    import re
                    email_match = re.search(r'[\w\.-]+@[\w\.-]+\.\w+', line)
                    if email_match:
                        data["Email"] = email_match.group()

        # 提取 box boxW listInner 里的字段
        box_div = soup.find("div", class_="box boxW listInner")
        if box_div:
            # 定义字段映射
            field_mapping = {
                "ATM ID": "ATM ID",
                "Agency": "Agency",
                "Category": "Category",
                "Close Date & Time": "Close Date & Time",
                "Publish Date": "Publish Date",
                "Location": "Location",
                "ATM Type": "ATM Type",
                "Multi Agency Access": "Multi Agency Access",
                "Panel Arrangement": "Panel Arrangement",
                "Multi-stage": "Multi-stage",
                "Multi-stage Criteria": "Multi-stage Criteria",
                "Description": "Description",
                "Other Instructions": "Other Instructions",
                "Conditions for Participation": "Conditions for Participation",
                "Timeframe for Delivery": "Timeframe for Delivery",
                "Address for Lodgement": "Address for Lodgement",
                "Addenda Available": "Addenda Available",
            }

            # 查找所有标签-值对
            # 通常结构是 label 在某个标签内，值在相邻标签
            all_text = box_div.get_text(separator="\n", strip=True)
            lines = all_text.split("\n")
            
            current_label = None
            for i, line in enumerate(lines):
                line = line.strip()
                if not line:
                    continue
                
                # 检查是否是字段名
                for field_name in field_mapping.keys():
                    if line == field_name or line.startswith(field_name + ":"):
                        current_label = field_name
                        # 如果行内包含值（格式: "字段名: 值"）
                        if ":" in line and line != field_name + ":":
                            value = line.split(":", 1)[1].strip()
                            # 排除不需要的内容
                            if "Show close time for other time zones" not in value:
                                data[field_name] = value
                        break
                else:
                    # 当前行不是字段名，可能是上一个字段的值
                    if current_label and current_label in field_mapping:
                        # 排除不需要的内容
                        if "Show close time for other time zones" not in line:
                            # 去除开头的冒号
                            if line.startswith(":"):
                                line = line[1:].strip()
                            if data[current_label]:
                                data[current_label] += " " + line
                            else:
                                data[current_label] = line

            # 更精确的解析：查找 dt/dd 或 label/value 结构
            dt_elements = box_div.find_all("dt")
            for dt in dt_elements:
                label = dt.get_text(strip=True)
                dd = dt.find_next_sibling("dd")
                if dd:
                    value = dd.get_text(strip=True)
                    # 排除不需要的内容
                    if "Show close time for other time zones" in value:
                        value = value.replace("Show close time for other time zones", "").strip()
                    # 去除开头的冒号
                    if value.startswith(":"):
                        value = value[1:].strip()
                    for field_name, mapped_name in field_mapping.items():
                        if field_name in label:
                            data[mapped_name] = value
                            break

            # 如果 dt/dd 结构不存在，尝试其他结构
            if not dt_elements:
                # 尝试查找所有 strong/b 标签作为标签
                for strong in box_div.find_all(["strong", "b"]):
                    label = strong.get_text(strip=True)
                    # 获取相邻内容作为值
                    next_elem = strong.next_sibling
                    value = ""
                    while next_elem:
                        if hasattr(next_elem, 'name'):
                            if next_elem.name in ["strong", "b", "br"]:
                                break
                            value += next_elem.get_text(strip=True) + " "
                        else:
                            value += str(next_elem).strip() + " "
                        next_elem = next_elem.next_sibling
                    
                    value = value.strip()
                    if "Show close time for other time zones" in value:
                        value = value.replace("Show close time for other time zones", "").strip()
                    # 去除开头的冒号
                    if value.startswith(":"):
                        value = value[1:].strip()
                    
                    for field_name, mapped_name in field_mapping.items():
                        if field_name in label:
                            data[mapped_name] = value
                            break

    except Exception as e:
        print(f"  [警告] 详情页解析出错 ({url}): {e}")
        return None

    return data


# ==================== 网站爬取1：搜索页面爬取 ====================

def crawl_search_page(max_pages, start_page, status_filter="full", tender_types=None):
    """
    网站爬取1：从搜索页面爬取 ATM 列表
    
    status_filter: 状态筛选
        "full": 保存所有数据
        "close": 仅保存包含Closed的数据
        "open": 仅保存不包含Closed的数据
    tender_types: 详情页类型筛选列表，如 ["Atm", "Son", "Cn"]，None 表示全部类型
        可选值：Atm, Son, Cn
    """
    import time

    print("\n" + "=" * 60)
    print("【网站爬取1】从搜索页面爬取 ATM 列表")
    print(f"目标网址: {SEARCH_URL}")
    status_desc = {"full": "全部状态", "close": "仅Closed状态", "open": "仅非Closed状态"}
    print(f"筛选条件: {status_desc.get(status_filter, '全部状态')}")
    if tender_types:
        print(f"详情页类型: {', '.join(tender_types)}")
    else:
        print("详情页类型: 全部类型（Atm, Son, Cn）")
    print("=" * 60)

    batch_data = []
    counter = 0

    # 计算当前年份
    current_year = datetime.now().year
    os.makedirs(SAVE_DIR, exist_ok=True)
    file_name = f"澳大利亚-{current_year}年当前采购项目.xlsx"
    file_path = os.path.join(SAVE_DIR, file_name)

    with sync_playwright() as p:
        print("\n[步骤1] 初始化浏览器...")
        browser = p.chromium.launch(
            headless=True,
            args=[
                "--no-sandbox",
                "--disable-setuid-sandbox",
                "--disable-dev-shm-usage",
                "--disable-blink-features=AutomationControlled",
            ],
        )
        context = browser.new_context(
            viewport=VIEWPORT,
            user_agent=USER_AGENT,
            bypass_csp=True,
        )
        # 反检测脚本：绕过 CloudFront 等 WAF 的自动化浏览器识别
        context.add_init_script("""
            Object.defineProperty(navigator, 'webdriver', {
                get: () => undefined
            });
            Object.defineProperty(navigator, 'languages', {
                get: () => ['en-US', 'en']
            });
            Object.defineProperty(navigator, 'plugins', {
                get: () => [1, 2, 3, 4, 5]
            });
            window.chrome = {
                runtime: {}
            };
        """)
        page = context.new_page()
        page.set_default_timeout(120000)
        page.set_default_navigation_timeout(120000)

        print("[步骤2] 访问目标页面...")
        # 构造带页码的URL（第1页也需要page=1参数）
        initial_url = f"{SEARCH_URL}&page={start_page}"
        print(f"  正在访问: {initial_url}")
        page.goto(initial_url, wait_until="domcontentloaded", timeout=120000)
        page.wait_for_load_state("load", timeout=60000)
        print(f"  页面加载完成，当前第 {start_page} 页")

        page_num = start_page
        end_page = None
        if max_pages != "full" and max_pages > 0:
            end_page = start_page + max_pages - 1

        while True:
            current_url = f"{SEARCH_URL}&page={page_num}"
            print(f"\n--- 正在处理第 {page_num} 页 ---")
            print(f"  当前页URL: {current_url}")

            # 等待页面加载完成
            try:
                page.wait_for_timeout(5000)  # 等待页面渲染
            except Exception as e:
                print(f"  [错误] 等待页面加载失败: {e}")
                break

            # 收集当前页所有详情页链接
            try:
                list_divs = page.query_selector_all(LIST_SELECTOR)
                
                print(f"  本页找到 {len(list_divs)} 个列表项")

                if not list_divs:
                    print("  未找到列表项，爬取结束")
                    break

                detail_urls = []
                
                for div in list_divs:
                    # 在列表项中查找链接
                    links = div.query_selector_all("a[href]")
                    for link in links:
                        href = link.get_attribute("href")
                        
                        if href:
                            if href.startswith("/"):
                                href = BASE_URL + href
                            elif not href.startswith("http"):
                                href = BASE_URL + "/" + href
                            
                            # 匹配详情页链接格式：/Atm/Show/, /Son/Show/, /Cn/Show/
                            # 根据 tender_types 筛选
                            is_atm = "/Atm/Show/" in href
                            is_son = "/Son/Show/" in href
                            is_cn = "/Cn/Show/" in href
                            
                            if is_atm or is_son or is_cn:
                                # 如果指定了类型筛选，按类型过滤
                                if tender_types is not None:
                                    if is_atm and "Atm" in tender_types:
                                        detail_urls.append(href)
                                    elif is_son and "Son" in tender_types:
                                        detail_urls.append(href)
                                    elif is_cn and "Cn" in tender_types:
                                        detail_urls.append(href)
                                else:
                                    # 未指定类型筛选，获取全部
                                    detail_urls.append(href)

                # 去重
                detail_urls = list(dict.fromkeys(detail_urls))
                print(f"  找到 {len(detail_urls)} 个详情页链接")

            except Exception as e:
                print(f"  [错误] 收集链接失败: {e}")
                break

            # 逐个请求详情页
            for idx, detail_url in enumerate(detail_urls):
                try:
                    print(f"  [{idx + 1}/{len(detail_urls)}] 正在抓取: {detail_url}")
                    data = extract_detail(detail_url, FLAT_HEADERS_1, status_filter=status_filter)
                    
                    if data:
                        batch_data.append(data)
                        counter += 1

                        if counter % 100 == 0:
                            print(f"  [进展] 已抓取 {counter} 条数据")

                    time.sleep(REQUEST_DELAY)

                except Exception as e:
                    print(f"  [警告] 抓取失败 ({detail_url}): {e}")
                    time.sleep(REQUEST_DELAY * 2)
                    continue

            # 检查是否达到最大页数
            if end_page is not None and page_num >= end_page:
                print(f"  已达页数限制，爬取结束。")
                break

            # 尝试翻页 - 使用 page.goto 直接导航避免 AJAX 局部刷新导致超时
            try:
                page.wait_for_timeout(3000)
                
                # 先检查是否还有下一页
                next_button = page.query_selector(NEXT_PAGE_SELECTOR)
                
                if not next_button:
                    print("  未找到下一页按钮，爬取结束。")
                    break

                # 检查按钮是否可点击
                is_disabled = next_button.evaluate("""el => {
                    if (el.getAttribute('aria-disabled') === 'true') return true;
                    if (el.classList.contains('disabled')) return true;
                    return false;
                }""")

                if is_disabled:
                    print("  下一页按钮已禁用，爬取结束。")
                    break

                # 直接导航到下一页URL，避免 AJAX 局部刷新导致 wait_for_load_state 超时
                page_num += 1
                next_page_url = f"{SEARCH_URL}&page={page_num}"
                print(f"  导航到下一页: {next_page_url}")
                page.goto(next_page_url, wait_until="domcontentloaded", timeout=120000)
                page.wait_for_timeout(3000)

            except Exception as e:
                print(f"  翻页失败或已到最后一页: {e}")
                break

        browser.close()

    # 保存数据
    if batch_data:
        save_to_excel(batch_data, file_path, FLAT_HEADERS_1)
        return file_path, len(batch_data)
    
    return None, 0


# ==================== 网站爬取2：RSS XML 解析 ====================

def crawl_rss_xml(status_filter="full"):
    """
    网站爬取2：从 RSS XML 解析并爬取采购项目
    
    status_filter: 状态筛选
        "full": 保存所有数据
        "close": 仅保存包含Closed的数据
        "open": 仅保存不包含Closed的数据
    """
    import time

    print("\n" + "=" * 60)
    print("【网站爬取2】从 RSS XML 解析采购项目")
    print(f"RSS 地址: {RSS_URL}")
    status_desc = {"full": "全部状态", "close": "仅Closed状态", "open": "仅非Closed状态"}
    print(f"筛选条件: {status_desc.get(status_filter, '全部状态')}")
    print("=" * 60)

    batch_data = []
    counter = 0

    # 保存文件名
    os.makedirs(SAVE_DIR, exist_ok=True)
    file_path = os.path.join(SAVE_DIR, "澳大利亚-RSS采购数据.xlsx")

    # 获取 RSS XML
    print("\n[步骤1] 获取 RSS XML...")
    try:
        response = http_session.get(RSS_URL, timeout=60)
        response.raise_for_status()
    except Exception as e:
        print(f"  [错误] 获取 RSS XML 失败: {e}")
        return None, 0

    # 解析 XML
    print("[步骤2] 解析 XML...")
    try:
        root = ET.fromstring(response.content)
    except Exception as e:
        print(f"  [错误] 解析 XML 失败: {e}")
        return None, 0

    # 提取所有 item 的 link
    items = root.findall(".//item")
    print(f"  找到 {len(items)} 个 item")

    detail_urls = []
    for item in items:
        link_elem = item.find("link")
        if link_elem is not None and link_elem.text:
            detail_urls.append(link_elem.text)

    print(f"  找到 {len(detail_urls)} 个详情页链接")

    # 逐个请求详情页
    print("[步骤3] 抓取详情页...")
    for idx, detail_url in enumerate(detail_urls):
        try:
            print(f"  [{idx + 1}/{len(detail_urls)}] 正在抓取: {detail_url}")
            data = extract_detail(detail_url, FLAT_HEADERS_2, status_filter=status_filter)
            
            if data:
                batch_data.append(data)
                counter += 1

                if counter % 100 == 0:
                    print(f"  [进展] 已抓取 {counter} 条数据")

            time.sleep(REQUEST_DELAY)

        except Exception as e:
            print(f"  [警告] 抓取失败 ({detail_url}): {e}")
            time.sleep(REQUEST_DELAY * 2)
            continue

    # 保存数据
    if batch_data:
        save_to_excel(batch_data, file_path, FLAT_HEADERS_2)
        return file_path, len(batch_data)
    
    return None, 0


# ==================== Gofile 上传 ====================

def create_gofile_account():
    """创建 Gofile 匿名账户，返回 token 和 rootFolderId"""
    resp = requests.post("https://api.gofile.io/accounts")
    data = resp.json()["data"]
    return data["token"], data["rootFolder"]


def upload_to_gofile_single(filepath, token, folder_id, max_retries=2):
    """上传单个文件到 Gofile，支持重试，返回上传结果"""
    import time
    
    filename = os.path.basename(filepath)
    filesize = os.path.getsize(filepath)
    if filesize >= 1024 * 1024:  # >= 1MB
        size_str = f"{filesize / 1024 / 1024:.1f} MB"
    else:
        size_str = f"{filesize / 1024:.1f} KB"
    print(f"  上传: {filename} ({size_str})")

    last_error = None
    for attempt in range(max_retries + 1):
        try:
            if attempt > 0:
                print(f"  第 {attempt + 1} 次重试上传...")

            # 获取最佳服务器
            resp = requests.get(f"https://api.gofile.io/servers", params={"token": token}, timeout=30)
            servers = resp.json()["data"]["servers"]
            na_servers = [s["name"] for s in servers if s["zone"] == "na"]
            server = na_servers[0] if na_servers else servers[0]["name"]

            # 上传
            upload_url = f"https://{server}.gofile.io/contents/uploadfile"
            start_time = time.time()
            with open(filepath, "rb") as f:
                files = {"file": (filename, f)}
                data = {"token": token, "folderId": folder_id}
                resp = requests.post(upload_url, files=files, data=data, timeout=1800)

            elapsed = time.time() - start_time
            result = resp.json()
            if result.get("status") != "ok":
                raise Exception(f"上传失败: {result}")

            download_page = result["data"]["downloadPage"]
            print(f"  上传成功! ({elapsed:.1f}s) 下载页: {download_page}")

            return {
                "filename": filename,
                "downloadPage": download_page,
                "fileId": result["data"]["id"],
                "size": filesize,
            }
        except Exception as e:
            last_error = e
            print(f"  上传失败(尝试 {attempt + 1}/{max_retries + 1}): {e}")
            if attempt < max_retries:
                time.sleep(3)

    raise Exception(f"上传失败，已重试 {max_retries} 次: {last_error}")


def set_gofile_folder_public(folder_id, token):
    """设置 Gofile 文件夹为公开访问"""
    resp = requests.put(
        f"https://api.gofile.io/contents/{folder_id}/update",
        headers={
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json",
        },
        json={"attribute": "public", "attributeValue": True},
    )
    return resp.json()


def upload_to_gofile(file_paths):
    """
    将多个文件上传到 Gofile 同一目录
    返回: (download_link, upload_results)
    """
    import time
    
    print("\n[步骤] 上传文件到 Gofile.io...")

    # 创建账户
    token, root_folder = create_gofile_account()
    print(f"  Gofile 账户已创建")

    # 按文件大小排序，小文件先上传
    files_with_size = [(fp, os.path.getsize(fp)) for fp in file_paths]
    files_sorted = sorted(files_with_size, key=lambda x: x[1])
    
    upload_results = []
    for filepath, filesize in files_sorted:
        try:
            result = upload_to_gofile_single(filepath, token, root_folder)
            upload_results.append(result)
        except Exception as e:
            print(f"  上传失败: {os.path.basename(filepath)} - {e}")

    # 关键：设置文件夹公开访问
    if upload_results and token:
        first_result = upload_results[0]
        # 从 downloadPage 提取 folder code
        # downloadPage 格式类似: https://gofile.io/d/foldercode
        folder_code = first_result["downloadPage"].split("/")[-1]
        
        print("\n  设置文件夹为公开访问...")
        # 需要获取 folder id，从 upload result 中获取
        # 实际上我们需要用 root_folder
        try:
            set_gofile_folder_public(root_folder, token)
            print("  文件夹已设为公开")
        except Exception as e:
            print(f"  设置公开失败: {e}")

    download_link = upload_results[0]["downloadPage"] if upload_results else None
    print(f"  下载链接: {download_link}")
    
    return download_link, upload_results


# ==================== 邮件发送 ====================

def send_email(sender, auth_code, receiver, download_link, file_info_list):
    """
    发送包含下载链接的邮件
    receiver: 单个邮箱字符串，或多个邮箱用分号分隔的字符串
    """
    # 解析多个收件人（用分号分隔）
    if ';' in receiver:
        receiver_list = [r.strip() for r in receiver.split(';') if r.strip()]
    else:
        receiver_list = [receiver.strip()]
    
    print(f"\n[步骤] 发送邮件到 {len(receiver_list)} 个收件人: {', '.join(receiver_list)}...")

    subject = "澳大利亚 AusTender 招标数据爬取结果 - Excel下载链接"

    # 文件清单
    file_rows_html = ""
    file_rows_text = ""
    for i, info in enumerate(file_info_list, 1):
        bg = "#f2f2f2" if i % 2 == 1 else "#ffffff"
        file_rows_html += f"""<tr style="background-color: {bg};">
<td style="border: 1px solid #ddd; padding: 10px;">{i}</td>
<td style="border: 1px solid #ddd; padding: 10px;">{info['filename']}</td>
<td style="border: 1px solid #ddd; padding: 10px;">{info['records']:,}</td></tr>
"""
        file_rows_text += f"{i}. {info['filename']} ({info['records']}条)\n"

    now_str = datetime.now().strftime('%Y-%m-%d')
    total_records = sum(f['records'] for f in file_info_list)

    html_body = f"""<html><body style="font-family: Microsoft YaHei, Arial, sans-serif; color: #333; line-height: 1.8; max-width: 700px; margin: 0 auto;">
<h2 style="color: #1a5276; border-bottom: 2px solid #2980b9; padding-bottom: 8px;">澳大利亚 AusTender 招标数据爬取结果</h2>
<div style="background-color: #eaf2f8; padding: 15px; border-radius: 5px; margin: 15px 0;">
<strong>爬取时间：</strong>{now_str}<br>
<strong>数据来源：</strong>https://www.tenders.gov.au<br>
<strong>总计：</strong>{len(file_info_list)}个Excel文件，{total_records:,}条招标记录
</div>
<h3>下载链接</h3>
<div style="background-color: #fff3cd; border: 1px solid #ffc107; padding: 20px; border-radius: 5px; text-align: center; margin: 15px 0;">
<a href="{download_link}" style="font-size: 18px; color: #2980b9; font-weight: bold; text-decoration: none;">点击下载全部文件</a>
<br><span style="color: #888; font-size: 13px;">{download_link}</span>
</div>
<h3>文件清单</h3>
<table style="border-collapse: collapse; width: 100%;">
<tr style="background-color: #2980b9; color: white;"><th style="padding: 10px; text-align: left;">序号</th><th style="padding: 10px; text-align: left;">文件名</th><th style="padding: 10px; text-align: left;">记录数</th></tr>
{file_rows_html}
</table>
<div style="background-color: #f8d7da; border: 1px solid #f5c6cb; padding: 15px; border-radius: 5px; margin: 20px 0;">
<strong style="color: #721c24;">提醒：Gofile.io 链接有效期有限，建议尽快下载。</strong>
</div>
<p style="color:#888; font-size:12px; margin-top:20px;">此邮件由自动化爬取系统发送，请勿直接回复。</p>
</body></html>"""

    text_body = f"""澳大利亚 AusTender 招标数据爬取结果

爬取时间：{now_str}
数据来源：https://www.tenders.gov.au
总计：{len(file_info_list)}个Excel文件，{total_records:,}条招标记录

下载链接：{download_link}

文件清单：
{file_rows_text}
提醒：Gofile.io 链接有效期有限，建议尽快下载。
"""

    msg = MIMEMultipart("alternative")
    msg["From"] = sender
    msg["To"] = ", ".join(receiver_list)  # 多个收件人用逗号分隔
    msg["Subject"] = subject
    msg.attach(MIMEText(text_body, "plain", "utf-8"))
    msg.attach(MIMEText(html_body, "html", "utf-8"))

    try:
        with smtplib.SMTP_SSL("smtp.qq.com", 465) as server:
            server.login(sender, auth_code)
            server.sendmail(sender, receiver_list, msg.as_string())
        print("  邮件发送成功!")
    except Exception as e:
        print(f"  邮件发送失败: {e}")


def send_email_with_attachments(sender, auth_code, receiver, files, file_records):
    """
    发送带附件的邮件（符合QQ邮箱规则）
    receiver: 单个邮箱字符串，或多个邮箱用分号分隔的字符串
    """
    # 解析多个收件人（用分号分隔）
    if ';' in receiver:
        receiver_list = [r.strip() for r in receiver.split(';') if r.strip()]
    else:
        receiver_list = [receiver.strip()]
    
    print(f"\n[步骤] 发送邮件到 {len(receiver_list)} 个收件人: {', '.join(receiver_list)}（带附件）...")

    subject = "澳大利亚 AusTender 招标数据爬取结果 - Excel附件"

    # 文件清单
    file_rows_text = ""
    total_size = 0
    for i, (filepath, filename, filesize) in enumerate(files, 1):
        total_size += filesize
        if filesize >= 1024 * 1024:  # >= 1MB
            size_str = f"{filesize/1024/1024:.1f} MB"
        else:
            size_str = f"{filesize/1024:.1f} KB"
        records = file_records.get(filename, 0)
        file_rows_text += f"{i}. {filename} ({size_str}, {records}条记录)\n"

    now_str = datetime.now().strftime('%Y-%m-%d')
    total_records = sum(file_records.values())

    html_body = f"""<html><body style="font-family: Microsoft YaHei, Arial, sans-serif; color: #333; line-height: 1.8; max-width: 700px; margin: 0 auto;">
<h2 style="color: #1a5276; border-bottom: 2px solid #2980b9; padding-bottom: 8px;">澳大利亚 AusTender 招标数据爬取结果</h2>
<div style="background-color: #eaf2f8; padding: 15px; border-radius: 5px; margin: 15px 0;">
<strong>爬取时间：</strong>{now_str}<br>
<strong>数据来源：</strong>https://www.tenders.gov.au<br>
<strong>总计：</strong>{len(files)}个Excel文件，{total_records:,}条招标记录（作为附件发送）
</div>
<h3>文件清单</h3>
<pre style="background-color: #f5f5f5; padding: 15px; border-radius: 5px; font-family: Consolas, Monaco, monospace;">{file_rows_text}</pre>
<div style="background-color: #d4edda; border: 1px solid #c3e6cb; padding: 15px; border-radius: 5px; margin: 20px 0;">
<strong style="color: #155724;">✓ 文件已作为附件发送，可直接下载。</strong>
</div>
<p style="color:#888; font-size:12px; margin-top:20px;">此邮件由自动化爬取系统发送，请勿直接回复。</p>
</body></html>"""

    text_body = f"""澳大利亚 AusTender 招标数据爬取结果

爬取时间：{now_str}
数据来源：https://www.tenders.gov.au
总计：{len(files)}个Excel文件，{total_records:,}条招标记录（作为附件发送）

文件清单：
{file_rows_text}
文件已作为附件发送，可直接下载。
"""

    msg = MIMEMultipart("mixed")
    msg["From"] = sender
    msg["To"] = ", ".join(receiver_list)  # 多个收件人用逗号分隔
    msg["Subject"] = subject

    msg.attach(MIMEText(text_body, "plain", "utf-8"))
    msg.attach(MIMEText(html_body, "html", "utf-8"))

    # 添加附件
    for filepath, filename, filesize in files:
        try:
            with open(filepath, "rb") as f:
                part = MIMEBase("application", "octet-stream")
                part.set_payload(f.read())
                encoders.encode_base64(part)
                # 正确设置附件文件名
                part.add_header('Content-Disposition', 'attachment', filename=('utf-8', '', filename))
                msg.attach(part)
            print(f"  已添加附件: {filename}")
        except Exception as e:
            print(f"  添加附件失败: {filename} - {e}")

    try:
        with smtplib.SMTP_SSL("smtp.qq.com", 465) as server:
            server.login(sender, auth_code)
            server.sendmail(sender, receiver_list, msg.as_string())
        print("  邮件发送成功（带附件）!")
        return True
    except Exception as e:
        # 尝试TLS方式
        try:
            server = smtplib.SMTP("smtp.qq.com", 587)
            server.starttls()
            server.login(sender, auth_code)
            server.sendmail(sender, receiver_list, msg.as_string())
            server.quit()
            print("  邮件发送成功（带附件，TLS）！")
            return True
        except Exception as e2:
            print(f"  邮件发送失败: {e2}")
            return False


# ==================== 主流程 ====================

def main(crawl_type=None, max_pages="full", start_page=1, sender=None, auth_code=None, receiver=None, search_status="full", rss_status="full", tender_types=None, attachments=False):
    """
    主流程

    crawl_type: 爬取类型
        None 或不传: 执行两种爬取
        1: 仅执行网站爬取1（搜索页面）
        2: 仅执行网站爬取2（RSS XML）
    max_pages: 最大爬取页数（仅对网站爬取1有效）
        "full": 爬取全部页
        数字: 爬取指定页数
    start_page: 起始页码（仅对网站爬取1有效）
    sender/auth_code/receiver: 邮件参数
    search_status: 搜索页面状态筛选 (full/close/open)
    rss_status: RSS爬取状态筛选 (full/close/open)
    tender_types: 详情页类型筛选列表，如 ["Atm", "Son", "Cn"]，None 表示全部类型
    """
    print("=" * 60)
    print("澳大利亚 AusTender 招标数据爬取")
    print(f"保存目录: {SAVE_DIR}")
    print("=" * 60)

    saved_files = []
    saved_records = []

    # 根据爬取类型执行
    if crawl_type is None or crawl_type == 1:
        file_path, record_count = crawl_search_page(max_pages, start_page, status_filter=search_status, tender_types=tender_types)
        if file_path:
            saved_files.append(file_path)
            saved_records.append(record_count)

    if crawl_type is None or crawl_type == 2:
        file_path, record_count = crawl_rss_xml(status_filter=rss_status)
        if file_path:
            saved_files.append(file_path)
            saved_records.append(record_count)

    # 汇总
    print("\n" + "=" * 60)
    total_records = sum(saved_records)
    print(f"爬取完成！共抓取 {total_records} 条数据，保存为 {len(saved_files)} 个文件")
    print(f"保存目录: {SAVE_DIR}")
    print("=" * 60)

    # 上传并发送邮件
    if sender and auth_code and receiver and saved_files:
        # 获取所有保存的文件及其大小
        files = []
        file_records_dict = {}
        for fp, rec_count in zip(saved_files, saved_records):
            filepath = fp
            filename = os.path.basename(fp)
            filesize = os.path.getsize(filepath)
            files.append((filepath, filename, filesize))
            file_records_dict[filename] = rec_count

        # 判断发送方式
        # - 不传 --attachments：强制走 Gofile 方式
        # - 传 --attachments：按现有规则判断
        total_size = sum(f[2] for f in files)

        if not attachments:
            # 强制走 Gofile 方式
            can_send_as_attachment = False
        else:
            # 检查是否符合QQ邮箱附件规则
            can_send_as_attachment = True

            # 检查单个文件大小
            for filepath, filename, filesize in files:
                if filesize > QQ_MAX_SINGLE_FILE_SIZE:
                    can_send_as_attachment = False
                    print(f"\n  ⚠️ 文件 {filename} 大小 {filesize/1024/1024:.1f}MB 超过20MB限制")
                    break

            # 检查总大小
            if can_send_as_attachment and total_size > QQ_MAX_TOTAL_SIZE:
                can_send_as_attachment = False
            print(f"\n  ⚠️ 文件总大小 {total_size/1024/1024:.1f}MB 超过50MB限制")

        if can_send_as_attachment:
            # 方式1：作为邮件附件发送
            print("\n[步骤] 发送邮件（带附件）...")
            print(f"  文件总大小: {total_size/1024/1024:.1f}MB，符合QQ邮箱附件规则")
            
            try:
                send_email_with_attachments(
                    sender=sender,
                    auth_code=auth_code,
                    receiver=receiver,
                    files=files,
                    file_records=file_records_dict,
                )
            except Exception as e:
                print(f"  邮件发送失败: {e}")

            # 汇总
            print(f"\n{'='*60}")
            print(f"完成! 共下载 {len(saved_files)} 个文件")
            print(f"文件已作为附件发送")
            print(f"保存目录: {SAVE_DIR}")

        else:
            # 方式2：上传到Gofile，邮件发送下载链接
            print("\n[步骤] 上传文件到 Gofile.io...")
            
            try:
                download_link, upload_results = upload_to_gofile(saved_files)

                if download_link and upload_results:
                    # 准备文件信息列表
                    file_info_list = []
                    for result in upload_results:
                        filename = result["filename"]
                        records = file_records_dict.get(filename, 0)
                        file_info_list.append({
                            "filename": filename,
                            "records": records,
                        })
                    
                    send_email(sender, auth_code, receiver, download_link, file_info_list)
                else:
                    print("\n[警告] Gofile 上传失败，跳过邮件发送")
            except Exception as e:
                print(f"\n[警告] Gofile 上传或邮件发送失败: {e}")
                download_link = None
                upload_results = []

            # 汇总
            print(f"\n{'='*60}")
            print(f"完成! 共下载 {len(saved_files)} 个文件")
            if 'upload_results' in locals() and upload_results:
                print(f"上传 {len(upload_results)} 个文件到 Gofile")
                link_str = download_link if 'download_link' in locals() and download_link else 'N/A'
                print(f"Gofile 链接: {link_str}")
            print(f"保存目录: {SAVE_DIR}")

    elif saved_files:
        print("\n[提示] 未提供邮件参数，跳过上传和邮件发送")
        print("  如需上传并发送邮件，请使用:")
        print("  python australia_tender.py --sender 发件邮箱 --auth-code 授权码 --receiver 收件邮箱")
        print("  多个收件邮箱用分号分隔，如: --receiver xxx@qq.com;yyy@live.com")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="澳大利亚 AusTender 招标数据爬取、上传与邮件通知"
    )
    parser.add_argument(
        "--sender", required=False, default=None,
        help="发件邮箱（QQ 邮箱）"
    )
    parser.add_argument(
        "--auth-code", required=False, default=None,
        help="发件邮箱 SMTP 授权码"
    )
    parser.add_argument(
        "--receiver", required=False, default=None,
        help="收件邮箱"
    )
    parser.add_argument(
        "--max-pages", required=False, default="full",
        help="最大爬取页数，默认 full（全部页）。可以是数字 n（爬取 n 页）或 'full'"
    )
    parser.add_argument(
        "--start-page", required=False, default="1",
        help="起始页码，默认 1"
    )
    parser.add_argument(
        "--type", required=False, default=None,
        help="爬取类型：不传或传0执行两种爬取，1 仅执行搜索页面爬取，2 仅执行 RSS XML 爬取"
    )
    parser.add_argument(
        "--tender-type", required=False, default=None,
        help="详情页类型筛选：不传则爬取全部类型，可指定 Atm/Son/Cn（多个用逗号分隔）"
    )
    parser.add_argument(
        "--search-status", required=False, default="full",
        help="搜索页面状态筛选：full=全部（默认）, close=仅Closed, open=仅非Closed"
    )
    parser.add_argument(
        "--rss-status", required=False, default="full",
        help="RSS爬取状态筛选：full=全部（默认）, close=仅Closed, open=仅非Closed"
    )
    parser.add_argument(
        "--attachments", action="store_true",
        help="传此参数时，文件符合规则则作为邮件附件发送；不传则强制走 Gofile 链接方式"
    )
    args = parser.parse_args()

    # 解析 max_pages 参数
    max_pages_arg = args.max_pages
    if max_pages_arg.lower() == "full":
        max_pages = "full"
    else:
        try:
            max_pages = int(max_pages_arg)
        except ValueError:
            print(f"警告: --max-pages 参数无效 '{args.max_pages}'，使用默认值 full")
            max_pages = "full"

    # 解析 start_page 参数
    try:
        start_page = int(args.start_page)
        if start_page < 1:
            print(f"警告: --start-page 参数无效 '{args.start_page}'，使用默认值 1")
            start_page = 1
    except ValueError:
        print(f"警告: --start-page 参数无效 '{args.start_page}'，使用默认值 1")
        start_page = 1

    # 解析 type 参数
    crawl_type = None
    if args.type:
        try:
            crawl_type = int(args.type)
            if crawl_type not in [0, 1, 2]:
                print(f"警告: --type 参数无效 '{args.type}'，应为 0、1 或 2，将执行两种爬取")
                crawl_type = None
            elif crawl_type == 0:
                crawl_type = None  # 0 表示执行两种爬取
        except ValueError:
            print(f"警告: --type 参数无效 '{args.type}'，应为 0、1 或 2，将执行两种爬取")
            crawl_type = None

    # 解析 tender-type 参数
    tender_types = None
    valid_types = {"Atm", "Son", "Cn"}
    if args.tender_type:
        types_str = args.tender_type.strip()
        if types_str:
            tender_types = [t.strip() for t in types_str.split(",")]
            # 验证类型有效性
            invalid_types = [t for t in tender_types if t not in valid_types]
            if invalid_types:
                print(f"警告: --tender-type 参数包含无效类型: {invalid_types}，有效值为: {sorted(valid_types)}")
                tender_types = [t for t in tender_types if t in valid_types]
                if not tender_types:
                    tender_types = None
            print(f"[信息] 将爬取以下类型: {tender_types}")

    # 解析 search-status 参数
    valid_status = {"full", "close", "open"}
    search_status = args.search_status.strip().lower() if args.search_status else "full"
    if search_status not in valid_status:
        print(f"警告: --search-status 参数无效 '{args.search_status}'，有效值为: {sorted(valid_status)}，使用默认值 full")
        search_status = "full"

    # 解析 rss-status 参数
    rss_status = args.rss_status.strip().lower() if args.rss_status else "full"
    if rss_status not in valid_status:
        print(f"警告: --rss-status 参数无效 '{args.rss_status}'，有效值为: {sorted(valid_status)}，使用默认值 full")
        rss_status = "full"

    main(
        crawl_type=crawl_type,
        max_pages=max_pages,
        start_page=start_page,
        sender=args.sender,
        auth_code=args.auth_code,
        receiver=args.receiver,
        search_status=search_status,
        rss_status=rss_status,
        tender_types=tender_types,
        attachments=args.attachments
    )
